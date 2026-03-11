from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from usuarios.decorators import acesso_permitido_apenas_para_filial
from .decorators import acesso_permitido_por_aba
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from .utils import importar_agendamentos, enviar_whatsapp_api, criar_agendamento_manual, editar_agendamento_manual, timezone_now, timezone_today
from .models import Agendamento, Transportadora, Motorista, GrupoUsuario
from .mensagens import enviar_notificacao_etapa
import io
import pandas as pd
import json
from django.utils import timezone as django_timezone  # Manter para compatibilidade se necessário
from datetime import datetime, timedelta
from django.db.models import Q, Count, Sum, F, Case, When, IntegerField
from django.views.decorators.http import require_http_methods
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
import logging
from django.core.paginator import Paginator
import requests
import openpyxl

logger = logging.getLogger(__name__)


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('portaria')
def portaria_agendamentos(request):
    """
    View principal da portaria - ATUALIZADA para o novo fluxo
    AGUARDANDO: ainda não passou pela portaria
    LIBERADOS: já foram liberados na portaria (próximas etapas)
    """
    try:
        hoje = timezone_now().date()

        # Aguardando liberação na portaria
        agendados = Agendamento.objects.filter(
            data_agendada=hoje,
            portaria_liberacao__isnull=True
        ).select_related('transportadora', 'motorista').order_by('horario_agendado')

        # Já liberados na portaria (vão para checklist, armazém, etc)
        liberados = Agendamento.objects.filter(
            data_agendada=hoje,
            portaria_liberacao__isnull=False
        ).select_related('transportadora', 'motorista', 'portaria_liberado_por')

        # ESTRATÉGIA NUCLEAR: Separação Explicita de Listas
        # Garante que NENHUMA regra de ordenação de banco ou chave complexa falhe
        
        todos_liberados = list(liberados)
        
        # Grupo 1: Sem Armazém (Topo)
        sem_armazem = [x for x in todos_liberados if not x.portaria_chegada_armazem]
        # Ordenar: Mais recentes de liberação primeiro
        sem_armazem.sort(key=lambda x: x.portaria_liberacao.timestamp() if x.portaria_liberacao else 0, reverse=True)
        
        # Grupo 2: Com Armazém (Base)
        com_armazem = [x for x in todos_liberados if x.portaria_chegada_armazem]
        # Ordenar: Data de armazém antiga primeiro (FIFO)
        com_armazem.sort(key=lambda x: x.portaria_chegada_armazem.timestamp())
        
        # Concatenar
        liberados_final = sem_armazem + com_armazem
        
        # LOG DE DEPURAÇÃO
        if liberados_final:
            logger.info("=== ORDEM FINAL FORÇADA ===")
            for i, item in enumerate(liberados_final[:5]):
                status = "SEM_ARM" if not item.portaria_chegada_armazem else "COM_ARM"
                status = "SEM_ARM" if not item.portaria_chegada_armazem else "COM_ARM"
                # logger.info(f"{i}: {item.motorista.nome} - {status} - {item.portaria_liberacao}")





        context = {
            'agendamentos': agendados,
            'liberados': liberados_final,
            'total_agendamentos': agendados.count(),
            'total_liberados': liberados.count(),
            'agora': timezone_now(),
        }

        # OTIMIZAÇÃO: Se for requisição AJAX (SmartUpdate), retorna apenas o partial
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
             return render(request, 'partials/_tabela_portaria.html', context)

        return render(request, 'portaria_rdn.html', context)

    except Exception as e:
        logger.error(f"Erro em portaria_agendamentos: {str(e)}")
        messages.error(request, 'Erro ao carregar dados da portaria.')
        return render(request, 'portaria_rdn.html', {
            'agendamentos': [], 'liberados': [], 'agora': timezone_now()
        })


@login_required
@require_http_methods(["POST"])
def confirmar_chegada(request):
    """
    NOVA FUNÇÃO: Libera o veículo na portaria (substitui a antiga confirmar_chegada)
    """
    agendamento_id = request.POST.get('agendamento_id')

    if not agendamento_id:
        return JsonResponse({'success': False, 'error': 'ID do agendamento não fornecido'})

    try:
        agendamento = get_object_or_404(Agendamento, id=agendamento_id)

        if agendamento.portaria_liberacao:
            return JsonResponse({'success': False, 'error': 'Veículo já foi liberado na portaria'})

        # Liberação na portaria
        # Usar timezone_now() que retorna datetime naive no fuso de Rondonópolis
        # Salvar diretamente no horário de Rondonópolis (sem conversão para UTC)
        agora_rdn = timezone_now()
        agendamento.portaria_liberacao = agora_rdn
        agendamento.portaria_liberado_por = request.user
        agendamento.atualizar_status_geral()
        agendamento.save()

        # Enviar notificações para o grupo checklist (assíncrono, não bloqueia resposta)
        enviar_notificacao_etapa(agendamento, 'portaria', request.user)

        logger.info(f"PORTARIA - Liberação confirmada por {request.user} - Placa: {agendamento.placa_veiculo}")

        # Forçar atualização explícita das telas afetadas (Garantia extra além dos signals)
        try:
             from .models import ControleAtualizacao
             from django.utils import timezone
             agora = timezone.now()
             ControleAtualizacao.objects.update_or_create(tela='portaria', defaults={'ultima_atualizacao': agora})
             ControleAtualizacao.objects.update_or_create(tela='checklist', defaults={'ultima_atualizacao': agora})
             ControleAtualizacao.objects.update_or_create(tela='agendamentos', defaults={'ultima_atualizacao': agora})
        except Exception as e:
             logger.error(f"Erro ao forçar atualização explícita (portaria): {e}")

        return JsonResponse({
            'success': True,
            'message': 'Veículo liberado na portaria! Agora aguarda CheckList.',
            'novo_status': agendamento.get_status_geral_display()
        })

    except Exception as e:
        logger.error(f"Erro em confirmar_chegada (liberar portaria): {str(e)}")
        return JsonResponse({'success': False, 'error': 'Erro ao liberar veículo'})


@login_required
@require_http_methods(["POST"])
def registrar_saida(request):
    """
    Mantida apenas para compatibilidade (pode ser removida depois)
    Agora não faz mais parte do fluxo principal
    """
    return JsonResponse({
        'success': False,
        'error': 'Função desativada no novo fluxo'
    })


@login_required
@require_http_methods(["GET"])
def detalhes_agendamento(request):
    """
    Retorna os detalhes completos do agendamento com todas as novas etapas e usuários
    """
    agendamento_id = request.GET.get('id')

    if not agendamento_id:
        return JsonResponse({'success': False, 'error': 'ID do agendamento não fornecido'})

    try:
        ag = get_object_or_404(Agendamento, id=agendamento_id)

        data = {
            'success': True,
            'agendamento': {
                'id': ag.id,
                'ordem': ag.ordem,
                'motorista': ag.motorista.nome,
                'horario_agendado': ag.horario_agendado.strftime('%H:%M'),
                'data_agendada': ag.data_agendada.strftime('%d/%m/%Y'),
                'tipo': ag.get_tipo_display(),
                'placa_veiculo': ag.placa_veiculo,
                'transportadora': ag.transportadora.nome,
                'peso': float(ag.peso),
                'tipo_veiculo': ag.get_tipo_veiculo_display(),
                'observacoes': ag.observacoes,
                'coluna_ad': ag.coluna_ad,
                'status_geral': ag.get_status_geral_display(),

                # Portaria
                'portaria_liberacao': ag.portaria_liberacao.strftime('%d/%m/%Y %H:%M') if ag.portaria_liberacao else None,
                'portaria_por': ag.portaria_liberado_por.get_full_name() if ag.portaria_liberado_por else None,

                # CheckList
                'checklist_numero': ag.checklist_numero,
                'checklist_data': ag.checklist_data.strftime('%d/%m/%Y %H:%M') if ag.checklist_data else None,
                'checklist_por': ag.checklist_preenchido_por.get_full_name() if ag.checklist_preenchido_por else None,
                'checklist_observacao': ag.checklist_observacao,

                # Armazém
                'armazem_chegada': ag.armazem_chegada.strftime('%d/%m/%Y %H:%M') if ag.armazem_chegada else None,
                'armazem_por': ag.armazem_confirmado_por.get_full_name() if ag.armazem_confirmado_por else None,

                # Onda
                'onda_status': ag.get_onda_status_display(),
                'onda_liberacao': ag.onda_liberacao.strftime('%d/%m/%Y %H:%M') if ag.onda_liberacao else None,
                'onda_por': ag.onda_liberado_por.get_full_name() if ag.onda_liberado_por else None,
            }
        }
        return JsonResponse(data)

    except Exception as e:
        logger.error(f"Erro em detalhes_agendamento: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Erro interno do servidor'})


@login_required
@require_http_methods(["POST"])
def confirmar_chegada_multipla(request):
    """
    Libera múltiplos veículos na portaria de uma vez
    """
    agendamentos_ids = request.POST.getlist('agendamentos_ids[]')

    if not agendamentos_ids:
        return JsonResponse({'success': False, 'error': 'Nenhum agendamento selecionado'})

    resultados = {'sucessos': 0, 'erros': 0, 'detalhes': []}

    for agendamento_id in agendamentos_ids:
        try:
            ag = Agendamento.objects.get(id=agendamento_id)

            if ag.portaria_liberacao:
                resultados['erros'] += 1
                resultados['detalhes'].append({
                    'id': agendamento_id, 'placa': ag.placa_veiculo,
                    'sucesso': False, 'erro': 'Já liberado na portaria'
                })
                continue

            # Usar timezone_now() que retorna datetime naive no fuso de Rondonópolis
            # Salvar diretamente no horário de Rondonópolis (sem conversão para UTC)
            agora_rdn = timezone_now()
            ag.portaria_liberacao = agora_rdn
            ag.portaria_liberado_por = request.user
            ag.atualizar_status_geral()
            ag.save()

            # Enviar notificações para o grupo checklist (assíncrono, não bloqueia resposta)
            enviar_notificacao_etapa(ag, 'portaria', request.user)

            resultados['sucessos'] += 1
            resultados['detalhes'].append({
                'id': agendamento_id, 'placa': ag.placa_veiculo, 'sucesso': True
            })

        except Agendamento.DoesNotExist:
            resultados['erros'] += 1
            resultados['detalhes'].append({'id': agendamento_id, 'sucesso': False, 'erro': 'Não encontrado'})
        except Exception as e:
            resultados['erros'] += 1
            resultados['detalhes'].append({'id': agendamento_id, 'sucesso': False, 'erro': str(e)})

    if resultados['sucessos'] > 0:
        # Forçar atualização explícita das telas afetadas (Garantia extra além dos signals)
        try:
             from .models import ControleAtualizacao
             from django.utils import timezone
             agora = timezone.now()
             ControleAtualizacao.objects.update_or_create(tela='portaria', defaults={'ultima_atualizacao': agora})
             ControleAtualizacao.objects.update_or_create(tela='checklist', defaults={'ultima_atualizacao': agora})
             ControleAtualizacao.objects.update_or_create(tela='agendamentos', defaults={'ultima_atualizacao': agora})
        except Exception as e:
             logger.error(f"Erro ao forçar atualização explícita (portaria multipla): {e}")

    return JsonResponse({
        'success': resultados['sucessos'] > 0,
        'resultados': resultados
    })


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('agendamentos')
def lista_agendamentos(request):
    try:
        agendamentos = Agendamento.objects.all().select_related('transportadora')
        transportadoras = Transportadora.objects.all()

        return render(request, 'lista_agendamentos.html', {
            'agendamentos': agendamentos,
            'transportadoras': transportadoras,
        })
    except Exception as e:
        logger.error(f"Erro na lista_agendamentos: {str(e)}")
        messages.error(request, 'Erro ao carregar lista.')
        return render(request, 'lista_agendamentos.html', {'agendamentos': [], 'transportadoras': []})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
def configuracoes_perfil(request):
    from .models import PreferenciaNotificacaoUsuario, ConfiguracaoNotificacao
    
    # Buscar ou criar preferências do usuário
    try:
        preferencias = PreferenciaNotificacaoUsuario.objects.get(usuario=request.user)
    except PreferenciaNotificacaoUsuario.DoesNotExist:
        preferencias = PreferenciaNotificacaoUsuario.objects.create(
            usuario=request.user,
            receber_email=True,
            receber_whatsapp=True,
            receber_navegador=True
        )
    
    # Buscar configuração de notificação (email/whatsapp configurados pelo admin)
    try:
        config = ConfiguracaoNotificacao.objects.get(usuario=request.user)
    except ConfiguracaoNotificacao.DoesNotExist:
        config = None
    
    # Se for POST, salvar apenas preferências (usuário não pode alterar email/telefone)
    if request.method == 'POST':
        receber_email = request.POST.get('receber_email') == 'on'
        receber_whatsapp = request.POST.get('receber_whatsapp') == 'on'
        receber_navegador = request.POST.get('receber_navegador') == 'on'
        
        preferencias.receber_email = receber_email
        preferencias.receber_whatsapp = receber_whatsapp
        preferencias.receber_navegador = receber_navegador
        preferencias.save()
        
        messages.success(request, 'Preferências de notificação salvas com sucesso!')
        return redirect('rondonopolis:configuracoes_perfil')
    
    return render(request, 'configeperfil.html', {
        'preferencias': preferencias,
        'config_notificacao': config,
    })


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
def importar_agendamentos_view(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Use POST'})

    arquivo = request.FILES.get('arquivo')
    if not arquivo:
        return JsonResponse({'success': False, 'error': 'Nenhum arquivo enviado'})

    if not arquivo.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({'success': False, 'error': 'Apenas arquivos Excel'})

    if arquivo.size > 10 * 1024 * 1024:
        return JsonResponse({'success': False, 'error': 'Arquivo muito grande (máx 10MB)'})

    try:
        # CORRETO — remove o parâmetro "usuario" que não existe na função
        resultado = importar_agendamentos(arquivo)
        return JsonResponse(resultado)
    except Exception as e:
        logger.error(f"Erro na importação: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erro: {str(e)}'})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
def criar_agendamento_view(request):
    """
    View para criar agendamento manualmente através do formulário
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Use POST'})
    
    try:
        # Processar dados do formulário
        dados = {
            'motorista': request.POST.get('motorista', ''),
            'data_agendada': request.POST.get('data_agendada', ''),
            'horario_agendado': request.POST.get('horario_agendado', ''),
            'tipo': request.POST.get('tipo', ''),
            'placa_veiculo': request.POST.get('placa_veiculo', ''),
            'transportadora': request.POST.get('transportadora', ''),
            'peso': request.POST.get('peso', '0'),
            'tipo_veiculo': request.POST.get('tipo_veiculo', ''),
            'observacoes': request.POST.get('observacoes', ''),
            'ordem': request.POST.get('ordem', ''),
            'encaixe': request.POST.get('encaixe', '')
        }
        
        resultado = criar_agendamento_manual(dados)
        return JsonResponse(resultado)
        
    except Exception as e:
        logger.error(f"Erro na criação de agendamento: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erro: {str(e)}'})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@require_http_methods(["GET"])
def obter_agendamento(request):
    """
    View para obter dados de um agendamento específico para edição
    """
    agendamento_id = request.GET.get('id')
    
    if not agendamento_id:
        return JsonResponse({'success': False, 'error': 'ID do agendamento não fornecido'})
    
    try:
        agendamento = get_object_or_404(Agendamento, id=agendamento_id)
        
        # Formatar data e hora para o formulário
        data_agendada_str = agendamento.data_agendada.strftime('%Y-%m-%d')
        horario_agendado_str = agendamento.horario_agendado.strftime('%H:%M')
        
        # Mapear tipo de veículo para maiúsculas (formato do formulário)
        tipo_veiculo_map = {
            'vuc': 'VUC',
            'toco': 'TOCO',
            'truck': 'TRUCK',
            'carreta': 'CARRETA',
            'bitrem': 'BITREM',
            'rodotrem': 'RODOTREM',
            'ls': 'LS'
        }
        tipo_veiculo_display = tipo_veiculo_map.get(agendamento.tipo_veiculo, agendamento.tipo_veiculo.upper())
        
        data = {
            'success': True,
            'agendamento': {
                'id': agendamento.id,
                'ordem': agendamento.ordem,
                'motorista': agendamento.motorista.nome,
                'motorista_id': agendamento.motorista.id,
                'data_agendada': data_agendada_str,
                'horario_agendado': horario_agendado_str,
                'tipo': agendamento.tipo,
                'placa_veiculo': agendamento.placa_veiculo,
                'transportadora_id': agendamento.transportadora.id,
                'transportadora_nome': agendamento.transportadora.nome,
                'peso': float(agendamento.peso),
                'tipo_veiculo': tipo_veiculo_display,
                'observacoes': agendamento.observacoes or '',
            }
        }
        return JsonResponse(data)
        
    except Exception as e:
        logger.error(f"Erro ao obter agendamento: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Erro ao buscar agendamento'})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@require_http_methods(["POST"])
def editar_agendamento_view(request):
    """
    View para editar agendamento através do formulário
    """
    agendamento_id = request.POST.get('agendamento_id')
    
    if not agendamento_id:
        return JsonResponse({'success': False, 'error': 'ID do agendamento não fornecido'})
    
    try:
        # Processar dados do formulário
        dados = {
            'motorista': request.POST.get('motorista', ''),
            'data_agendada': request.POST.get('data_agendada', ''),
            'horario_agendado': request.POST.get('horario_agendado', ''),
            'tipo': request.POST.get('tipo', ''),
            'placa_veiculo': request.POST.get('placa_veiculo', ''),
            'transportadora': request.POST.get('transportadora', ''),
            'peso': request.POST.get('peso', '0'),
            'tipo_veiculo': request.POST.get('tipo_veiculo', ''),
            'observacoes': request.POST.get('observacoes', ''),
            'encaixe': request.POST.get('encaixe', '')
        }
        
        resultado = editar_agendamento_manual(agendamento_id, dados)
        return JsonResponse(resultado)
        
    except Exception as e:
        logger.error(f"Erro na edição de agendamento: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erro: {str(e)}'})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('agendamentos')
@require_http_methods(["POST"])
def excluir_agendamento_view(request, agendamento_id):
    """
    View para excluir um agendamento
    """
    try:
        agendamento = get_object_or_404(Agendamento, id=agendamento_id)
        agendamento_id_backup = agendamento.id
        
        # Enviar atualização via WebSocket antes de excluir
        try:
            from .websocket_utils import enviar_atualizacao_portaria
            enviar_atualizacao_portaria('deleted', agendamento=agendamento)
        except Exception as e:
            logger.error(f"Erro ao enviar atualização WebSocket: {str(e)}")
            # Não falhar a exclusão se a atualização WebSocket falhar
        
        agendamento.delete()
        
        logger.info(f"Agendamento {agendamento_id_backup} excluído por {request.user}")
        
        return JsonResponse({
            'success': True,
            'message': 'Agendamento excluído com sucesso!'
        })
        
    except Exception as e:
        logger.error(f"Erro ao excluir agendamento {agendamento_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Erro ao excluir agendamento: {str(e)}'
        })


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
def baixar_modelo_importacao(request):
    colunas = ['ORDEM', 'MOTORISTA', 'DATA AGENDAMENTO', 'TIPO', 'TRANSPORTADORA', 'PLACA', 'VEICULO', 'PESO']
    exemplo = [{
        'ORDEM': 'IMP001', 'MOTORISTA': 'João Silva', 'DATA AGENDAMENTO': '20/12/2025',
        'TIPO': 'entrega', 'TRANSPORTADORA': 'Trans ABC', 'PLACA': 'ABC1D23',
        'VEICULO': 'carreta', 'PESO': '28000'
    }]

    df = pd.DataFrame(exemplo, columns=colunas)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Modelo')
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modelo_importacao.xlsx"'
    return response


@require_http_methods(["GET"])
def documento_impressao(request):
    """
    Gera HTML do documento de internamento (agora sem restrição de saída)
    """
    agendamento_id = request.GET.get('id')
    if not agendamento_id:
        return JsonResponse({'success': False, 'error': 'ID não informado'})

    try:
        agendamento = get_object_or_404(Agendamento, id=agendamento_id)
        # Aplicar a mesma lógica de ordenação das etapas do template de acompanhamento
        etapas_ordenadas = get_etapas_ordenadas(agendamento)
        html = render_to_string('documento_internamento.html', {
            'agendamento': agendamento,
            'etapas_ordenadas': etapas_ordenadas
        })
        return JsonResponse({'success': True, 'html': html})
    except Exception as e:
        logger.error(f"Erro no documento_impressao: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Erro ao gerar documento'})





@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('checklist')
@require_http_methods(["GET"])
def checklist_atualizar_dados(request):
    """
    View para buscar dados atualizados do Checklist via AJAX
    """
    try:
        hoje = timezone_now().date()

        # Pendentes: liberados na portaria mas sem checklist (APENAS COLETA)
        pendentes = Agendamento.objects.filter(
            data_agendada=hoje,
            portaria_liberacao__isnull=False,
            checklist_data__isnull=True,
            tipo='coleta'
        ).select_related('transportadora', 'motorista', 'portaria_liberado_por').order_by('portaria_liberacao')

        # Concluídos: já preencheram checklist (APENAS COLETA)
        concluidos = Agendamento.objects.filter(
            data_agendada=hoje,
            checklist_data__isnull=False,
            tipo='coleta'
        ).select_related('transportadora', 'motorista', 'checklist_preenchido_por').order_by('-checklist_data')

        # Serializar pendentes
        pendentes_data = []
        for agendamento in pendentes:
            pendentes_data.append({
                'id': agendamento.id,
                'motorista': agendamento.motorista.nome,
                'placa_veiculo': agendamento.placa_veiculo,
                'tipo': agendamento.tipo,
                'transportadora': agendamento.transportadora.nome,
                'tipo_veiculo': agendamento.get_tipo_veiculo_display(),
                'portaria_liberacao': agendamento.portaria_liberacao.isoformat() if agendamento.portaria_liberacao else None,
                'portaria_liberado_por': agendamento.portaria_liberado_por.get_full_name() if agendamento.portaria_liberado_por else None,
            })

        # Serializar concluídos
        concluidos_data = []
        for concluido in concluidos:
            concluidos_data.append({
                'id': concluido.id,
                'motorista': concluido.motorista.nome,
                'placa_veiculo': concluido.placa_veiculo,
                'checklist_numero': concluido.checklist_numero or '',
                'checklist_data': concluido.checklist_data.isoformat() if concluido.checklist_data else None,
                'checklist_preenchido_por': concluido.checklist_preenchido_por.get_full_name() if concluido.checklist_preenchido_por else None,
                'checklist_observacao': concluido.checklist_observacao or '',
            })

        return JsonResponse({
            'success': True,
            'pendentes': pendentes_data,
            'concluidos': concluidos_data,
            'total_pendentes': len(pendentes_data),
            'total_concluidos': len(concluidos_data),
            'timestamp': timezone_now().isoformat(),
        })

    except Exception as e:
        logger.error(f"Erro ao atualizar dados do checklist: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Erro ao atualizar dados: {str(e)}'
        })

# ... todo o seu código original permanece exatamente como está ...
# (portaria_agendamentos, confirmar_chegada, detalhes_agendamento, etc.)

# ==========================================================
# APENAS ADICIONE ESTAS DUAS FUNÇÕES NO FINAL DO SEU views.py
# ==========================================================

@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('checklist')
def checklist(request):
    """
    Tela do CheckList - mostra veículos liberados na portaria e aguardando checklist
    """
    hoje = timezone_now().date()

    pendentes = Agendamento.objects.filter(
        data_agendada=hoje,
        portaria_liberacao__isnull=False,
        checklist_data__isnull=True,
        tipo='coleta'
    ).select_related('transportadora', 'motorista', 'portaria_liberado_por').order_by('portaria_liberacao')

    concluidos = Agendamento.objects.filter(
        data_agendada=hoje,
        checklist_data__isnull=False
    ).select_related('transportadora', 'motorista', 'checklist_preenchido_por').order_by('-checklist_data')

    return render(request, 'checklist.html', {
        'pendentes': pendentes,
        'concluidos': concluidos,
        'total_pendentes': pendentes.count(),
        'total_concluidos': concluidos.count(),
    })

@login_required
@require_http_methods(["POST"])
def preencher_checklist(request):
    """
    Preenche o CheckList (apenas o número é obrigatório)
    Aceita ação individual ou em massa
    """
    # Suporta seleção múltipla ou ação individual
    ids = request.POST.getlist('agendamentos_ids[]')
    if not ids:
        single_id = request.POST.get('agendamento_id')
        if single_id:
            ids = [single_id]

    numero_checklist = request.POST.get('numero_checklist', '').strip()
    observacao = request.POST.get('observacao', '').strip()  # ← opcional

    if not ids:
        return JsonResponse({'success': False, 'error': 'Nenhum veículo selecionado'})

    if not numero_checklist:
        return JsonResponse({'success': False, 'error': 'Número do CheckList é obrigatório'})

    sucessos = 0
    erros = 0

    for agendamento_id in ids:
        try:
            ag = Agendamento.objects.select_for_update().get(id=agendamento_id)

            # Evita preenchimento duplicado
            if ag.checklist_data is not None:
                erros += 1
                continue

            ag.checklist_numero = numero_checklist
            # Usar timezone_now() que retorna datetime naive no fuso de Rondonópolis
            # Salvar diretamente no horário de Rondonópolis (sem conversão para UTC)
            agora_rdn = timezone_now()
            ag.checklist_data = agora_rdn
            ag.checklist_preenchido_por = request.user
            ag.checklist_observacao = observacao if observacao else None  # ← pode ser vazio
            ag.atualizar_status_geral()
            ag.save()

            # Enviar atualização via WebSocket para todas as telas afetadas
            try:
                enviar_atualizacao_tela('checklist', 'updated', agendamento=ag)
                enviar_atualizacao_tela('armazem', 'updated', agendamento=ag)
            except Exception as e:
                logger.error(f"Erro ao enviar atualização WebSocket: {str(e)}")

            # Enviar notificações para a próxima etapa (Armazém)
            enviar_notificacao_etapa(ag, 'checklist', request.user)

            sucessos += 1
            logger.info(f"CHECKLIST #{numero_checklist} - {request.user} - Placa: {ag.placa_veiculo}")

        except Agendamento.DoesNotExist:
            erros += 1
        except Exception as e:
            logger.error(f"Erro checklist ID {agendamento_id}: {e}")
            erros += 1

    if sucessos > 0:
        # Forçar atualização explícita das telas afetadas (Garantia extra além dos signals)
        try:
             from .models import ControleAtualizacao
             agora = timezone_now()
             ControleAtualizacao.objects.update_or_create(tela='checklist', defaults={'ultima_atualizacao': agora})
             ControleAtualizacao.objects.update_or_create(tela='onda', defaults={'ultima_atualizacao': agora})
             ControleAtualizacao.objects.update_or_create(tela='armazem', defaults={'ultima_atualizacao': agora})
        except Exception as e:
             logger.error(f"Erro ao forçar atualização explícita (checklist): {e}")

        return JsonResponse({
            'success': True,
            'message': f'CheckList #{numero_checklist} registrado em {sucessos} veículo(s)!'
        })
    else:
        return JsonResponse({
            'success': False,
            'error': 'Nenhum CheckList foi registrado.'
        })


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('armazem')
def armazem(request):
    hoje = timezone_now().date()

    # ARMAZÉM: Lógica combinada
    # COLETA: Precisa de CHECKLIST
    # ENTREGA: Basta ter passado na PORTARIA
    
    # PENDENTES: Itens que ainda não finalizaram a operação (aguardando OU em operação)
    # Inclui: armazem_chegada NULL (aguardando) + armazem_chegada NOT NULL mas armazem_saida NULL (em operação)
    pendentes = Agendamento.objects.filter(
        data_agendada=hoje,
        armazem_saida__isnull=True  # Não finalizou a operação
    ).filter(
        # Caso Coleta: ter checklist
        (Q(tipo='coleta') & Q(checklist_data__isnull=False)) |
        # Caso Entrega: ter portaria (checklist é null)
        (Q(tipo='entrega') & Q(portaria_liberacao__isnull=False))
    ).select_related('transportadora', 'motorista').order_by('horario_agendado')

    # CONCLUÍDOS: Veículos que finalizaram a operação do armazém (saída registrada)
    concluidos = Agendamento.objects.filter(
        data_agendada=hoje,
        armazem_saida__isnull=False  # Finalizou a operação
    ).select_related('transportadora', 'motorista', 'armazem_confirmado_por', 'armazem_saida_por').order_by('-armazem_saida')

    return render(request, 'armazem.html', {
        'pendentes': pendentes,
        'concluidos': concluidos,
        'total_pendentes': pendentes.count(),
        'total_concluidos': concluidos.count(),
    })


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('armazem')
@require_http_methods(["GET"])
def armazem_atualizar_dados(request):
    """
    View para buscar dados atualizados do Armazém via AJAX
    """
    try:
        hoje = timezone_now().date()

        # PENDENTES: Itens que ainda não finalizaram a operação (aguardando OU em operação)
        pendentes = Agendamento.objects.filter(
            data_agendada=hoje,
            armazem_saida__isnull=True  # Não finalizou a operação
        ).filter(
            # Caso Coleta: ter checklist
            (Q(tipo='coleta') & Q(checklist_data__isnull=False)) |
            # Caso Entrega: ter portaria (checklist é null)
            (Q(tipo='entrega') & Q(portaria_liberacao__isnull=False))
        ).select_related('transportadora', 'motorista').order_by('horario_agendado')

        # CONCLUÍDOS: Veículos que finalizaram a operação do armazém
        concluidos = Agendamento.objects.filter(
            data_agendada=hoje,
            armazem_saida__isnull=False  # Finalizou a operação
        ).select_related('transportadora', 'motorista', 'armazem_confirmado_por', 'armazem_saida_por').order_by('-armazem_saida')

        # Serializar pendentes
        pendentes_data = []
        for agendamento in pendentes:
            pendentes_data.append({
                'id': agendamento.id,
                'motorista': agendamento.motorista.nome,
                'placa_veiculo': agendamento.placa_veiculo,
                'tipo': agendamento.tipo,
                'transportadora': agendamento.transportadora.nome,
                'onda_liberacao': agendamento.onda_liberacao.isoformat() if agendamento.onda_liberacao else None,
                'checklist_data': agendamento.checklist_data.isoformat() if agendamento.checklist_data else None,
                'armazem_chegada': agendamento.armazem_chegada.isoformat() if agendamento.armazem_chegada else None,
            })

        # Serializar concluídos
        concluidos_data = []
        for concluido in concluidos:
            concluidos_data.append({
                'id': concluido.id,
                'motorista': concluido.motorista.nome,
                'placa_veiculo': concluido.placa_veiculo,
                'tipo': concluido.tipo,
                'transportadora': concluido.transportadora.nome,
                'armazem_chegada': concluido.armazem_chegada.isoformat() if concluido.armazem_chegada else None,
                'armazem_saida': concluido.armazem_saida.isoformat() if concluido.armazem_saida else None,
                'armazem_confirmado_por': concluido.armazem_confirmado_por.get_full_name() if concluido.armazem_confirmado_por else None,
                'armazem_saida_por': concluido.armazem_saida_por.get_full_name() if concluido.armazem_saida_por else None,
            })

        return JsonResponse({
            'success': True,
            'pendentes': pendentes_data,
            'concluidos': concluidos_data,
            'total_pendentes': len(pendentes_data),
            'total_concluidos': len(concluidos_data),
            'timestamp': timezone_now().isoformat(),
        })

    except Exception as e:
        logger.error(f"Erro ao atualizar dados do armazém: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Erro ao atualizar dados: {str(e)}'
        })


@login_required
@require_http_methods(["POST"])
def armazem_registrar_entrada(request):
    """
    Registra a chegada do veículo no armazém (após CheckList)
    OBRIGATÓRIO: A onda deve estar liberada antes de confirmar o armazém
    Usa os campos já existentes: armazem_chegada e armazem_confirmado_por
    """
    ids = request.POST.getlist('ids[]')  # vem do swipe múltiplo
    if not ids:
        return JsonResponse({'success': False, 'error': 'Nenhum veículo selecionado'})

    sucessos = 0
    erros_onda = []
    for ag_id in ids:
        try:
            with transaction.atomic():
                ag = Agendamento.objects.select_for_update().get(id=ag_id)

                # Já entrou? Pula
                if ag.armazem_chegada is not None:
                    continue

                # VALIDAÇÃO CRÍTICA: Verificar se a onda está liberada (APENAS SE FOR COLETA)
                if ag.tipo == 'coleta' and ag.onda_liberacao is None:
                    erros_onda.append(ag.placa_veiculo)
                    continue

                # Registra entrada
                # Usar timezone_now() que retorna datetime naive no fuso de Rondonópolis
                # Salvar diretamente no horário de Rondonópolis (sem conversão para UTC)
                agora_rdn = timezone_now()
                ag.armazem_chegada = agora_rdn
                ag.armazem_confirmado_por = request.user
                ag.atualizar_status_geral()  # vai mudar para 'pendente_liberacao_documentos'
                ag.save()

                # Enviar atualização via WebSocket para todas as telas afetadas
                try:
                    from .websocket_utils import enviar_atualizacao_tela
                    enviar_atualizacao_tela('armazem', 'updated', agendamento=ag)
                    enviar_atualizacao_tela('documentos', 'updated', agendamento=ag)
                except Exception as e:
                    logger.error(f"Erro ao enviar atualização WebSocket: {str(e)}")

                # Enviar notificações para o grupo de documentos
                # SOLICITAÇÃO 12/01: Documentos só deve receber notificação na SAÍDA do armazém.
                # enviar_notificacao_etapa(ag, 'armazem', request.user)

                sucessos += 1
                logger.info(f"ARMAZÉM - Entrada registrada por {request.user} - Placa: {ag.placa_veiculo}")

        except Agendamento.DoesNotExist:
            continue
        except Exception as e:
            logger.error(f"Erro ao registrar entrada no armazém (ID {ag_id}): {e}")

    if sucessos > 0:
        # Forçar atualização explícita das telas afetadas (Garantia extra além dos signals)
        try:
             from .models import ControleAtualizacao
             agora = timezone_now()
             ControleAtualizacao.objects.update_or_create(tela='armazem', defaults={'ultima_atualizacao': agora})
             ControleAtualizacao.objects.update_or_create(tela='liberacao-documentos', defaults={'ultima_atualizacao': agora})
        except Exception as e:
             logger.error(f"Erro ao forçar atualização explícita (armazem): {e}")

    if sucessos > 0 and not erros_onda:
        return JsonResponse({
            'success': True,
            'message': f'Entrada no armazém registrada em {sucessos} veículo(s)!'
        })
    elif sucessos > 0 and erros_onda:
        return JsonResponse({
            'success': True,
            'message': f'Entrada registrada em {sucessos} veículo(s). {len(erros_onda)} veículo(s) não puderam ser confirmados: onda não liberada (placas: {", ".join(erros_onda)})'
        })
    elif erros_onda:
        return JsonResponse({
            'success': False,
            'error': f'Não é possível confirmar o armazém: a onda deve estar liberada primeiro. Placas: {", ".join(erros_onda)}'
        })
    else:
        return JsonResponse({
            'success': False,
            'error': 'Nenhum veículo foi registrado (já estavam no armazém ou erro)'
        })

# ==============================================================================================


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('onda')
def liberacao_onda(request):
    """
    View para liberação de onda - pode ser liberada em qualquer momento do processo
    Sempre mostra apenas os agendamentos do dia atual
    """
    hoje = timezone_now().date()
    
    # Processar filtro de data
    data_filtro = request.GET.get('data')
    if data_filtro:
        try:
            data_selecionada = datetime.strptime(data_filtro, '%Y-%m-%d').date()
        except ValueError:
            data_selecionada = hoje
    else:
        data_selecionada = hoje
    
    # Query base - filtra pela data selecionada
    # REMOVIDO filtro de tipo='coleta' para incluir Entregas (OD)
    pendentes = Agendamento.objects.filter(
        data_agendada=data_selecionada,
        onda_liberacao__isnull=True
    ).select_related('transportadora', 'motorista').order_by('horario_agendado', 'data_agendada')

    concluidos = Agendamento.objects.filter(
        data_agendada=data_selecionada,
        onda_liberacao__isnull=False
    ).select_related('transportadora', 'motorista', 'onda_liberado_por').order_by('-onda_liberacao')

    return render(request, 'liberacao_onda.html', {
        'pendentes': pendentes,
        'concluidos': concluidos,
        'agora': timezone_now(),
        'total_pendentes': pendentes.count(),
        'total_concluidos': concluidos.count(),
        'data_selecionada': data_selecionada.strftime('%Y-%m-%d'),
    })


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('onda')
@require_http_methods(["GET"])
def onda_atualizar_dados(request):
    """
    View para buscar dados atualizados da Liberação de Onda via AJAX
    """
    try:
        hoje = timezone_now().date()
        
        # Processar filtro de data
        data_filtro = request.GET.get('data')
        if data_filtro:
            try:
                data_selecionada = datetime.strptime(data_filtro, '%Y-%m-%d').date()
            except ValueError:
                data_selecionada = hoje
        else:
            data_selecionada = hoje
        
        # Pendentes: agendamentos sem liberação de onda
        pendentes = Agendamento.objects.filter(
            data_agendada=data_selecionada,
            onda_liberacao__isnull=True
        ).select_related('transportadora', 'motorista').order_by('horario_agendado', 'data_agendada')

        # Concluídos: já liberados
        concluidos = Agendamento.objects.filter(
            data_agendada=data_selecionada,
            onda_liberacao__isnull=False
        ).select_related('transportadora', 'motorista').order_by('-onda_liberacao')

        # Serializar pendentes
        pendentes_data = []
        for agendamento in pendentes:
            pendentes_data.append({
                'id': agendamento.id,
                'motorista': agendamento.motorista.nome,
                'placa_veiculo': agendamento.placa_veiculo,
                'tipo': agendamento.tipo,
                'transportadora': agendamento.transportadora.nome,
                'tipo_veiculo': agendamento.get_tipo_veiculo_display(),
                'horario_agendado': agendamento.horario_agendado.strftime('%H:%M') if agendamento.horario_agendado else '00:00',
                'peso': str(agendamento.peso),
                'data_agendada': agendamento.data_agendada.isoformat(),
                'coluna_ad': agendamento.coluna_ad or '',
                'status_geral': agendamento.status_geral,
            })

        # Serializar concluídos
        concluidos_data = []
        for concluido in concluidos:
            concluidos_data.append({
                'id': concluido.id,
                'motorista': concluido.motorista.nome,
                'placa_veiculo': concluido.placa_veiculo,
                'tipo': concluido.tipo,
                'transportadora': concluido.transportadora.nome,
                'tipo_veiculo': concluido.get_tipo_veiculo_display(),
                'onda_liberacao': concluido.onda_liberacao.isoformat() if concluido.onda_liberacao else None,
                'onda_liberado_por': concluido.onda_liberado_por.get_full_name() if concluido.onda_liberado_por else None,
                'coluna_ad': concluido.coluna_ad or '',
                'data_agendada': concluido.data_agendada.isoformat(),
                'horario_agendado': concluido.horario_agendado.strftime('%H:%M') if concluido.horario_agendado else '00:00',
            })

        return JsonResponse({
            'success': True,
            'pendentes': pendentes_data,
            'concluidos': concluidos_data,
            'total_pendentes': len(pendentes_data),
            'total_concluidos': len(concluidos_data),
            'timestamp': timezone_now().isoformat(),
        })

    except Exception as e:
        logger.error(f"Erro ao atualizar dados da liberação de onda: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Erro ao atualizar dados: {str(e)}'
        })


@login_required
@require_http_methods(["POST"])
def onda_registrar_liberacao(request):
    """
    Registra a liberação da onda (após armazém)
    Usa os campos já existentes: onda_liberacao e onda_liberado_por
    """
    ids = request.POST.getlist('ids[]')  # vem do swipe múltiplo
    if not ids:
        return JsonResponse({'success': False, 'error': 'Nenhum veículo selecionado'})

    sucessos = 0
    for ag_id in ids:
        try:
            with transaction.atomic():
                ag = Agendamento.objects.select_for_update().get(id=ag_id)

                # Já foi liberado? Pula
                if ag.onda_liberacao is not None:
                    continue

                # Registra liberação
                # Usar timezone_now() que retorna datetime naive no fuso de Rondonópolis
                # Salvar diretamente no horário de Rondonópolis (sem conversão para UTC)
                agora_rdn = timezone_now()
                ag.onda_liberacao = agora_rdn
                ag.onda_liberado_por = request.user
                ag.onda_status = 'liberado'
                ag.atualizar_status_geral()  # vai mudar para 'processo_concluido'
                ag.save()

                # Enviar atualização via WebSocket para todas as telas afetadas
                try:
                    enviar_atualizacao_tela('onda', 'updated', agendamento=ag)
                    enviar_atualizacao_tela('armazem', 'updated', agendamento=ag)
                except Exception as e:
                    logger.error(f"Erro ao enviar atualização WebSocket: {str(e)}")

                # Enviar notificações para a próxima etapa (Armazém ou Documentos dependendo do fluxo)
                # SOLICITAÇÃO 12/01: Não enviar email individual na liberação da Onda.
                # A etapa apenas aparecerá como concluída nos emails das etapas subsequentes.
                # enviar_notificacao_etapa(ag, 'onda', request.user)

                sucessos += 1
                logger.info(f"ONDA - Liberação registrada por {request.user} - Placa: {ag.placa_veiculo}")

        except Agendamento.DoesNotExist:
            continue
        except Exception as e:
            logger.error(f"Erro ao registrar liberação da onda (ID {ag_id}): {e}")

    if sucessos > 0:
        # Forçar atualização explícita das telas afetadas (Garantia extra além dos signals)
        try:
             from .models import ControleAtualizacao
             agora = timezone_now()
             ControleAtualizacao.objects.update_or_create(tela='onda', defaults={'ultima_atualizacao': agora})
             ControleAtualizacao.objects.update_or_create(tela='armazem', defaults={'ultima_atualizacao': agora})
        except Exception as e:
             logger.error(f"Erro ao forçar atualização explícita (onda): {e}")

        return JsonResponse({
            'success': True,
            'message': f'Liberação da onda registrada em {sucessos} veículo(s)!'
        })
    else:
        return JsonResponse({
            'success': False,
            'error': 'Nenhum veículo foi liberado (já estavam liberados ou erro)'
        })


# ==============================================================================================


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('liberacao_documentos')
def liberacao_documentos(request):
    """
    View para documentos - última etapa do processo
    Só mostra agendamentos que já passaram pelo armazém
    Sempre mostra apenas os agendamentos do dia atual
    """
    hoje = timezone_now().date()
    
    # Query base - apenas agendamentos que entraram no armazém e ainda não tiveram documentos liberados
    pendentes = Agendamento.objects.filter(
        data_agendada=hoje,
        armazem_chegada__isnull=False,
        documentos_liberacao__isnull=True
    ).select_related('transportadora', 'motorista').order_by('armazem_chegada')

    concluidos = Agendamento.objects.filter(
        data_agendada=hoje,
        documentos_liberacao__isnull=False
    ).select_related('transportadora', 'motorista', 'documentos_liberado_por').order_by('-documentos_liberacao')

    return render(request, 'liberacao_documentos.html', {
        'pendentes': pendentes,
        'concluidos': concluidos,
        'agora': timezone_now(),
        'total_pendentes': pendentes.count(),
        'total_concluidos': concluidos.count(),
    })


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('liberacao_documentos')
@require_http_methods(["GET"])
def documentos_atualizar_dados(request):
    """
    View para buscar dados atualizados de Documentos via AJAX
    """
    try:
        hoje = timezone_now().date()
        
        # Pendentes: já entraram no armazém mas ainda não tiveram documentos liberados
        pendentes = Agendamento.objects.filter(
            data_agendada=hoje,
            armazem_chegada__isnull=False,
            documentos_liberacao__isnull=True
        ).select_related('transportadora', 'motorista', 'armazem_confirmado_por').order_by('armazem_chegada')

        # Concluídos: já tiveram documentos liberados
        concluidos = Agendamento.objects.filter(
            data_agendada=hoje,
            documentos_liberacao__isnull=False
        ).select_related('transportadora', 'motorista', 'documentos_liberado_por').order_by('-documentos_liberacao')

        # Serializar pendentes
        pendentes_data = []
        for agendamento in pendentes:
            pendentes_data.append({
                'id': agendamento.id,
                'motorista': agendamento.motorista.nome,
                'placa_veiculo': agendamento.placa_veiculo,
                'tipo': agendamento.tipo,
                'transportadora': agendamento.transportadora.nome,
                'data_agendada': agendamento.data_agendada.isoformat(),
                'horario_agendado': agendamento.horario_agendado.strftime('%H:%M') if agendamento.horario_agendado else '00:00',
                'coluna_ad': agendamento.coluna_ad or '',
                'armazem_chegada': agendamento.armazem_chegada.isoformat() if agendamento.armazem_chegada else None,
                'armazem_confirmado_por': agendamento.armazem_confirmado_por.get_full_name() if agendamento.armazem_confirmado_por else None,
            })

        # Serializar concluídos
        concluidos_data = []
        for concluido in concluidos:
            concluidos_data.append({
                'id': concluido.id,
                'motorista': concluido.motorista.nome,
                'placa_veiculo': concluido.placa_veiculo,
                'tipo': concluido.tipo,
                'transportadora': concluido.transportadora.nome,
                'documentos_liberacao': concluido.documentos_liberacao.isoformat() if concluido.documentos_liberacao else None,
                'documentos_liberado_por': concluido.documentos_liberado_por.get_full_name() if concluido.documentos_liberado_por else None,
            })

        return JsonResponse({
            'success': True,
            'pendentes': pendentes_data,
            'concluidos': concluidos_data,
            'total_pendentes': len(pendentes_data),
            'total_concluidos': len(concluidos_data),
            'timestamp': timezone_now().isoformat(),
        })

    except Exception as e:
        logger.error(f"Erro ao atualizar dados de documentos: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Erro ao atualizar dados: {str(e)}'
        })


@login_required
@require_http_methods(["POST"])
def documentos_registrar_liberacao(request):
    """
    Registra a liberação dos documentos (última etapa)
    Usa os campos: documentos_liberacao e documentos_liberado_por
    """
    ids = request.POST.getlist('ids[]')
    if not ids:
        return JsonResponse({'success': False, 'error': 'Nenhum veículo selecionado'})

    sucessos = 0
    for ag_id in ids:
        try:
            with transaction.atomic():
                ag = Agendamento.objects.select_for_update().get(id=ag_id)

                # Já foi liberado? Pula
                if ag.documentos_liberacao is not None:
                    continue
                
                # Verificar se passou pelo armazém
                if ag.armazem_chegada is None:
                    continue

                # Registra liberação
                # Usar timezone_now() que retorna datetime naive no fuso de Rondonópolis
                agora_rdn = timezone_now()
                ag.documentos_liberacao = agora_rdn
                ag.documentos_liberado_por = request.user
                
                # Salvar observação se fornecida
                observacao = request.POST.get('observacao', '').strip()
                if observacao:
                    ag.documentos_observacao = observacao
                
                ag.atualizar_status_geral()  # vai mudar para 'processo_concluido'
                ag.save()

                # Enviar atualização via WebSocket para todas as telas afetadas
                try:
                    from .websocket_utils import enviar_atualizacao_tela
                    enviar_atualizacao_tela('documentos', 'updated', agendamento=ag)
                except Exception as e:
                    logger.error(f"Erro ao enviar atualização WebSocket: {str(e)}")

                # Enviar notificações para todos os grupos quando processo for concluído
                enviar_notificacao_etapa(ag, 'documentos', request.user)

                sucessos += 1
                logger.info(f"DOCUMENTOS - Liberação registrada por {request.user} - Placa: {ag.placa_veiculo}")

        except Agendamento.DoesNotExist:
            continue
        except Exception as e:
            logger.error(f"Erro ao registrar liberação dos documentos (ID {ag_id}): {e}")

    if sucessos > 0:
        # Forçar atualização explícita das telas afetadas (Garantia extra além dos signals)
        try:
             from .models import ControleAtualizacao
             agora = timezone_now()
             ControleAtualizacao.objects.update_or_create(tela='liberacao-documentos', defaults={'ultima_atualizacao': agora})
        except Exception as e:
             logger.error(f"Erro ao forçar atualização explícita (documentos): {e}")

        return JsonResponse({
            'success': True,
            'message': f'Documentos liberados em {sucessos} veículo(s)!'
        })
    else:
        return JsonResponse({
            'success': False,
            'error': 'Nenhum veículo foi liberado (já estavam liberados ou erro)'
        })


# ==============================================================================================


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('processos')
def visualizacao_processos(request):
    """
    View para visualização completa de todos os processos
    """
    try:
        # Obter todos os agendamentos inicialmente
        # Otimização: select_related completo para evitar queries N+1
        agendamentos_list = Agendamento.objects.select_related(
            'transportadora', 'motorista', 'portaria_liberado_por',
            'checklist_preenchido_por', 'armazem_confirmado_por', 
            'documentos_liberado_por', 'criado_por'
        ).all()
        
        # Aplicar filtros baseados nos parâmetros GET
        status_filter = request.GET.get('status')
        tipo_filter = request.GET.get('tipo')
        data_filter = request.GET.get('data')
        busca_filter = request.GET.get('busca')
        
        # Filtro por status
        if status_filter:
            agendamentos_list = agendamentos_list.filter(status_geral=status_filter)
        
        # Filtro por tipo (coleta/entrega)
        if tipo_filter:
            agendamentos_list = agendamentos_list.filter(tipo=tipo_filter)
        
        # Filtro por data
        if data_filter:
            try:
                data_obj = datetime.strptime(data_filter, '%Y-%m-%d').date()
                agendamentos_list = agendamentos_list.filter(data_agendada=data_obj)
            except ValueError:
                pass
        
        # Filtro por busca (ordem, motorista, placa, transportadora)
        if busca_filter:
            agendamentos_list = agendamentos_list.filter(
                Q(ordem__icontains=busca_filter) |
                Q(motorista__icontains=busca_filter) |
                Q(placa_veiculo__icontains=busca_filter) |
                Q(transportadora__nome__icontains=busca_filter)
            )
        
        # Ordenação padrão
        agendamentos_list = agendamentos_list.order_by('-data_agendada', 'horario_agendado')
        
        # Processar cada agendamento para adicionar etapas ordenadas
        # Processar cada agendamento para adicionar etapas ordenadas
        agendamentos_com_etapas = []
        for agendamento in agendamentos_list:
            etapas = get_etapas_ordenadas(agendamento)
            
            # Calcular progresso
            total_etapas = len(etapas)
            etapas_concluidas = sum(1 for e in etapas if e['concluida'])
            
            agendamento_dict = {
                'obj': agendamento,
                'etapas_ordenadas': etapas,
                'total_etapas': total_etapas,
                'etapas_concluidas': etapas_concluidas
            }
            agendamentos_com_etapas.append(agendamento_dict)
        
        # Paginação - 10 itens por página
        paginator = Paginator(agendamentos_com_etapas, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'agendamentos': page_obj,
        }
        
        return render(request, 'visualizacao_processos.html', context)
        
    except Exception as e:
        logger.error(f"Erro em visualizacao_processos: {str(e)}")
        messages.error(request, 'Erro ao carregar visualização de processos.')
        return render(request, 'visualizacao_processos.html', {'agendamentos': []})





from django.shortcuts import render
from django.db.models import Count, Q, Sum
from .models import Agendamento, Transportadora
from datetime import datetime
from .utils import timezone_now, timezone_today

@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('painel')
def processos_painel(request):
    # Filtros
    data_filtro = request.GET.get('data')
    status_filtro = request.GET.get('status')
    tipo_filtro = request.GET.get('tipo')
    transportadora_filtro = request.GET.get('transportadora')
    
    # Data padrão: hoje
    if data_filtro:
        try:
            data_selecionada = datetime.strptime(data_filtro, '%Y-%m-%d').date()
        except ValueError:
            data_selecionada = timezone_now().date()
    else:
        data_selecionada = timezone_now().date()
    
    # Query base
    agendamentos = Agendamento.objects.filter(data_agendada=data_selecionada)
    
    # Aplicar filtros adicionais
    if status_filtro:
        agendamentos = agendamentos.filter(status_geral=status_filtro)
    
    if tipo_filtro:
        agendamentos = agendamentos.filter(tipo=tipo_filtro)
    
    if transportadora_filtro:
        agendamentos = agendamentos.filter(transportadora_id=transportadora_filtro)
    
    # Contadores por status
    contadores = agendamentos.aggregate(
        total=Count('id'),
        aguardando=Count('id', filter=Q(status_geral='aguardando_chegada')),
        checklist=Count('id', filter=Q(status_geral='em_checklist')),
        armazem=Count('id', filter=Q(status_geral='confirmacao_armazem')),
        pendente=Count('id', filter=Q(status_geral='pendente_liberacao_onda')),
        concluido=Count('id', filter=Q(status_geral='processo_concluido'))
    )
    
    # Transportadoras para o filtro
    transportadoras = Transportadora.objects.all()
    
    context = {
        'agendamentos': agendamentos.select_related('transportadora', 'portaria_liberado_por', 
                                                   'checklist_preenchido_por', 'armazem_confirmado_por',
                                                   'onda_liberado_por').order_by('horario_agendado'),
        'hoje': data_selecionada,
        'transportadoras': transportadoras,
        'total_count': contadores['total'] or 0,
        'aguardando_count': contadores['aguardando'] or 0,
        'checklist_count': contadores['checklist'] or 0,
        'armazem_count': contadores['armazem'] or 0,
        'pendente_count': contadores['pendente'] or 0,
        'concluido_count': contadores['concluido'] or 0,
    }
    
    return render(request, 'processos_painel.html', context)
from django.contrib.auth.decorators import login_required
from usuarios.decorators import acesso_permitido_apenas_para_filial
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from .utils import importar_agendamentos, enviar_whatsapp_api, criar_agendamento_manual, timezone_now, timezone_today
from .models import Agendamento, Transportadora, Motorista
import io
import pandas as pd
from datetime import datetime
from django.db.models import Q
from django.views.decorators.http import require_http_methods
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
import logging
from django.core.paginator import Paginator
import requests
from django.db import transaction

logger = logging.getLogger(__name__)


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('portaria')
def portaria_agendamentos(request):
    """
    View principal da portaria - ATUALIZADA para o novo fluxo
    AGUARDANDO: ainda não passou pela portaria
    LIBERADOS: já foram liberados na portaria (próximas etapas)
    """
    try:
        hoje = timezone_now().date()

        # Aguardando liberação na portaria
        agendados = Agendamento.objects.filter(
            data_agendada=hoje,
            portaria_liberacao__isnull=True
        ).select_related('transportadora', 'motorista').order_by('horario_agendado')

        # Já liberados na portaria (vão para checklist, armazém, etc)
        liberados = Agendamento.objects.filter(
            data_agendada=hoje,
            portaria_liberacao__isnull=False
        ).select_related('transportadora', 'motorista', 'portaria_liberado_por').order_by('-portaria_liberacao')

        # Motoristas para a aba de telefones
        motoristas = Motorista.objects.all().order_by('nome')

        context = {
            'agendamentos': agendados,
            'liberados': liberados,
            'motoristas': motoristas,
            'total_agendamentos': agendados.count(),
            'total_liberados': liberados.count(),
            'agora': timezone_now(),
        }

        return render(request, 'portaria_rdn.html', context)

    except Exception as e:
        logger.error(f"Erro em portaria_agendamentos: {str(e)}")
        messages.error(request, 'Erro ao carregar dados da portaria.')
        return render(request, 'portaria_rdn.html', {
            'agendamentos': [], 'liberados': [], 'motoristas': [], 'agora': timezone_now()
        })


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('portaria')
@require_http_methods(["GET"])
def portaria_atualizar_dados(request):
    """
    View para buscar dados atualizados da portaria via AJAX
    """
    try:
        hoje = timezone_now().date()

        # Aguardando liberação na portaria
        agendados = Agendamento.objects.filter(
            data_agendada=hoje,
            portaria_liberacao__isnull=True
        ).select_related('transportadora', 'motorista').order_by('horario_agendado')

        # Já liberados na portaria
        liberados = Agendamento.objects.filter(
            data_agendada=hoje,
            portaria_liberacao__isnull=False
        ).select_related('transportadora', 'motorista', 'portaria_liberado_por')
        
        # ESTRATÉGIA NUCLEAR: Separação Explicita de Listas
        todos_liberados = list(liberados)
        
        sem_armazem = [x for x in todos_liberados if not x.portaria_chegada_armazem]
        sem_armazem.sort(key=lambda x: x.portaria_liberacao.timestamp() if x.portaria_liberacao else 0, reverse=True)
        
        com_armazem = [x for x in todos_liberados if x.portaria_chegada_armazem]
        com_armazem.sort(key=lambda x: x.portaria_chegada_armazem.timestamp())
        
        liberados = sem_armazem + com_armazem


        # Serializar agendamentos
        agendamentos_data = []
        for agendamento in agendados:
            agendamentos_data.append({
                'id': agendamento.id,
                'motorista': agendamento.motorista.nome,
                'motorista_telefone': agendamento.motorista.telefone if agendamento.motorista and agendamento.motorista.telefone else None,
                'horario_agendado': agendamento.horario_agendado.strftime('%H:%M') if agendamento.horario_agendado else '00:00',
                'placa_veiculo': agendamento.placa_veiculo,
                'transportadora': agendamento.transportadora.nome,
                'tipo_veiculo': agendamento.get_tipo_veiculo_display(),
                'tipo': agendamento.tipo,
                'peso': str(agendamento.peso),
                'coluna_ad': agendamento.coluna_ad or '',
            })

        liberados_data = []
        for liberado in liberados:
            liberados_data.append({
                'id': liberado.id,
                'motorista': liberado.motorista.nome,
                'horario_agendado': liberado.horario_agendado.strftime('%H:%M') if liberado.horario_agendado else '00:00',
                'placa_veiculo': liberado.placa_veiculo,
                'transportadora': liberado.transportadora.nome,
                'tipo_veiculo': liberado.get_tipo_veiculo_display(),
                'tipo': liberado.tipo,
                'portaria_liberacao': liberado.portaria_liberacao.isoformat() if liberado.portaria_liberacao else None,
                'portaria_chegada_armazem': liberado.portaria_chegada_armazem.isoformat() if liberado.portaria_chegada_armazem else None,
                'coluna_ad': liberado.coluna_ad or '',
            })

        return JsonResponse({
            'success': True,
            'agendamentos': agendamentos_data,
            'liberados': liberados_data,
            'total_agendamentos': len(agendamentos_data),
            'total_liberados': len(liberados_data),
            'timestamp': timezone_now().isoformat(),
        })

    except Exception as e:
        logger.error(f"Erro ao atualizar dados da portaria: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Erro ao atualizar dados: {str(e)}'
        })


@login_required
@require_http_methods(["POST"])
def confirmar_chegada(request):
    """
    NOVA FUNÇÃO: Libera o veículo na portaria (substitui a antiga confirmar_chegada)
    """
    agendamento_id = request.POST.get('agendamento_id')

    if not agendamento_id:
        return JsonResponse({'success': False, 'error': 'ID do agendamento não fornecido'})

    try:
        agendamento = get_object_or_404(Agendamento, id=agendamento_id)

        if agendamento.portaria_liberacao:
            return JsonResponse({'success': False, 'error': 'Veículo já foi liberado na portaria'})

        # Liberação na portaria
        # Usar timezone_now() que retorna datetime naive no fuso de Rondonópolis
        # Salvar diretamente no horário de Rondonópolis (sem conversão para UTC)
        agora_rdn = timezone_now()
        agendamento.portaria_liberacao = agora_rdn
        agendamento.portaria_liberado_por = request.user
        agendamento.atualizar_status_geral()
        agendamento.save()

        # Enviar notificações para o grupo checklist (assíncrono, não bloqueia resposta)
        enviar_notificacao_etapa(agendamento, 'portaria', request.user)

        logger.info(f"PORTARIA - Liberação confirmada por {request.user} - Placa: {agendamento.placa_veiculo}")

        return JsonResponse({
            'success': True,
            'message': 'Veículo liberado na portaria! Agora aguarda CheckList.',
            'novo_status': agendamento.get_status_geral_display()
        })

    except Exception as e:
        logger.error(f"Erro em confirmar_chegada (liberar portaria): {str(e)}")
        return JsonResponse({'success': False, 'error': 'Erro ao liberar veículo'})


@login_required
@require_http_methods(["GET"])
def obter_horario_atual(request):
    """
    Retorna o horário atual do servidor no fuso horário de Rondonópolis
    """
    from .utils import TIMEZONE_RONDONOPOLIS
    import pytz
    
    # Obter horário atual em UTC
    agora_utc = django_timezone.now()
    
    # Converter para Rondonópolis
    agora_rdn = agora_utc.astimezone(TIMEZONE_RONDONOPOLIS)
    
    # Criar timestamp ISO para o JavaScript interpretar corretamente
    # O JavaScript precisa saber que é Rondonópolis
    timestamp_iso = agora_rdn.isoformat()
    
    return JsonResponse({
        'success': True,
        'horario': agora_rdn.strftime('%H:%M:%S'),
        'horario_sem_segundos': agora_rdn.strftime('%H:%M'),
        'data': agora_rdn.strftime('%Y-%m-%d'),
        'timestamp': timestamp_iso,
        'timezone': 'America/Cuiaba'
    })


# Desabilitado temporariamente - funcionalidade de atualização automática
# @login_required
# @acesso_permitido_apenas_para_filial('rondonopolis')
# @require_http_methods(["GET"])
# def portaria_atualizar_dados(request):
#     """
#     Retorna os dados atualizados da portaria em JSON para atualização automática
#     """
#     try:
#         hoje = timezone_now().date()
#
#         # Aguardando liberação na portaria
#         agendados = Agendamento.objects.filter(
#             data_agendada=hoje,
#             portaria_liberacao__isnull=True
#         ).select_related('transportadora', 'motorista').order_by('horario_agendado')
#
#         # Já liberados na portaria
#         liberados = Agendamento.objects.filter(
#             data_agendada=hoje,
#             portaria_liberacao__isnull=False
#         ).select_related('transportadora', 'motorista').order_by('-portaria_liberacao')
#
#         # Serializar agendados
#         agendados_data = []
#         for ag in agendados:
#             agendados_data.append({
#                 'id': str(ag.id),
#                 'motorista': ag.motorista.nome if ag.motorista else '',
#                 'telefone_motorista': ag.motorista.telefone if ag.motorista and ag.motorista.telefone else '',
#                 'horario_agendado': ag.horario_agendado.strftime('%H:%M') if ag.horario_agendado else '',
#                 'horario_chegada': ag.horario_chegada.strftime('%H:%M') if ag.horario_chegada else None,
#                 'tipo': ag.tipo,
#                 'placa': ag.placa_veiculo or '',
#                 'transportadora': ag.transportadora.nome if ag.transportadora else '',
#                 'peso': ag.peso or 0,
#                 'tipo_veiculo': ag.get_tipo_veiculo_display(),
#             })
#
#         # Serializar liberados
#         liberados_data = []
#         for lib in liberados:
#             liberados_data.append({
#                 'id': str(lib.id),
#                 'motorista': lib.motorista.nome if lib.motorista else '',
#                 'horario_agendado': lib.horario_agendado.strftime('%H:%M') if lib.horario_agendado else '',
#                 'portaria_liberacao': lib.portaria_liberacao.strftime('%d/%m/%Y %H:%M') if lib.portaria_liberacao else None,
#                 'tipo': lib.tipo,
#                 'placa': lib.placa_veiculo or '',
#                 'transportadora': lib.transportadora.nome if lib.transportadora else '',
#                 'peso': lib.peso or 0,
#             })
#
#         return JsonResponse({
#             'success': True,
#             'total_agendamentos': agendados.count(),
#             'total_liberados': liberados.count(),
#             'agendamentos': agendados_data,
#             'liberados': liberados_data,
#         })
#
#     except Exception as e:
#         logger.error(f"Erro em portaria_atualizar_dados: {str(e)}")
#         return JsonResponse({
#             'success': False,
#             'error': 'Erro ao atualizar dados'
#         })


@login_required
@require_http_methods(["GET"])
def detalhes_agendamento(request):
    """
    Retorna os detalhes completos do agendamento com todas as novas etapas e usuários
    """
    agendamento_id = request.GET.get('id')

    if not agendamento_id:
        return JsonResponse({'success': False, 'error': 'ID do agendamento não fornecido'})

    try:
        ag = get_object_or_404(Agendamento, id=agendamento_id)

        data = {
            'success': True,
            'agendamento': {
                'id': ag.id,
                'ordem': ag.ordem,
                'motorista': ag.motorista.nome,
                'horario_agendado': ag.horario_agendado.strftime('%H:%M'),
                'data_agendada': ag.data_agendada.strftime('%d/%m/%Y'),
                'tipo': ag.get_tipo_display(),
                'placa_veiculo': ag.placa_veiculo,
                'transportadora': ag.transportadora.nome,
                'peso': float(ag.peso),
                'tipo_veiculo': ag.get_tipo_veiculo_display(),
                'observacoes': ag.observacoes,
                'coluna_ad': ag.coluna_ad,
                'status_geral': ag.get_status_geral_display(),

                # Portaria
                'portaria_liberacao': ag.portaria_liberacao.strftime('%d/%m/%Y %H:%M') if ag.portaria_liberacao else None,
                'portaria_por': ag.portaria_liberado_por.get_full_name() if ag.portaria_liberado_por else None,

                # CheckList
                'checklist_numero': ag.checklist_numero,
                'checklist_data': ag.checklist_data.strftime('%d/%m/%Y %H:%M') if ag.checklist_data else None,
                'checklist_por': ag.checklist_preenchido_por.get_full_name() if ag.checklist_preenchido_por else None,
                'checklist_observacao': ag.checklist_observacao,

                # Armazém
                'armazem_chegada': ag.armazem_chegada.strftime('%d/%m/%Y %H:%M') if ag.armazem_chegada else None,
                'armazem_por': ag.armazem_confirmado_por.get_full_name() if ag.armazem_confirmado_por else None,

                # Onda
                'onda_status': ag.get_onda_status_display(),
                'onda_liberacao': ag.onda_liberacao.strftime('%d/%m/%Y %H:%M') if ag.onda_liberacao else None,
                'onda_por': ag.onda_liberado_por.get_full_name() if ag.onda_liberado_por else None,
            }
        }
        return JsonResponse(data)

    except Exception as e:
        logger.error(f"Erro em detalhes_agendamento: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Erro interno do servidor'})


@login_required
@require_http_methods(["POST"])
def confirmar_chegada_multipla(request):
    """
    Libera múltiplos veículos na portaria de uma vez
    """
    agendamentos_ids = request.POST.getlist('agendamentos_ids[]')

    if not agendamentos_ids:
        return JsonResponse({'success': False, 'error': 'Nenhum agendamento selecionado'})

    resultados = {'sucessos': 0, 'erros': 0, 'detalhes': []}

    for agendamento_id in agendamentos_ids:
        try:
            ag = Agendamento.objects.get(id=agendamento_id)

            if ag.portaria_liberacao:
                resultados['erros'] += 1
                resultados['detalhes'].append({
                    'id': agendamento_id, 'placa': ag.placa_veiculo,
                    'sucesso': False, 'erro': 'Já liberado na portaria'
                })
                continue

            # Usar timezone_now() que retorna datetime naive no fuso de Rondonópolis
            # Salvar diretamente no horário de Rondonópolis (sem conversão para UTC)
            agora_rdn = timezone_now()
            ag.portaria_liberacao = agora_rdn
            ag.portaria_liberado_por = request.user
            ag.atualizar_status_geral()
            ag.save()

            # Enviar notificações para o grupo checklist (assíncrono, não bloqueia resposta)
            enviar_notificacao_etapa(ag, 'portaria', request.user)

            resultados['sucessos'] += 1
            resultados['detalhes'].append({
                'id': agendamento_id, 'placa': ag.placa_veiculo, 'sucesso': True
            })

        except Agendamento.DoesNotExist:
            resultados['erros'] += 1
            resultados['detalhes'].append({'id': agendamento_id, 'sucesso': False, 'erro': 'Não encontrado'})
        except Exception as e:
            resultados['erros'] += 1
            resultados['detalhes'].append({'id': agendamento_id, 'sucesso': False, 'erro': str(e)})

    return JsonResponse({
        'success': resultados['sucessos'] > 0,
        'resultados': resultados
    })


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('agendamentos')
def lista_agendamentos(request):
    try:
        # Query base
        agendamentos = Agendamento.objects.all().select_related('transportadora', 'motorista')
        
        # Aplicar filtros baseados nos parâmetros GET
        tipo_filter = request.GET.get('tipo')
        transportadora_filter = request.GET.get('transportadora')
        data_filter = request.GET.get('data')
        busca_filter = request.GET.get('busca')
        status_filter = request.GET.get('status')
        
        # Se não houver data no filtro, usar a data atual como padrão
        if not data_filter:
            data_filter = timezone_today().strftime('%Y-%m-%d')
        
        # Filtro por tipo (coleta/entrega)
        if tipo_filter:
            agendamentos = agendamentos.filter(tipo=tipo_filter)
        
        # Filtro por transportadora
        if transportadora_filter:
            agendamentos = agendamentos.filter(transportadora_id=transportadora_filter)
        
        # Filtro por data (sempre aplicado, padrão é hoje)
        try:
            data_obj = datetime.strptime(data_filter, '%Y-%m-%d').date()
            agendamentos = agendamentos.filter(data_agendada=data_obj)
        except ValueError:
            # Se a data for inválida, usar a data atual
            data_obj = timezone_today()
            agendamentos = agendamentos.filter(data_agendada=data_obj)
            data_filter = data_obj.strftime('%Y-%m-%d')
        
        # Filtro por busca (ordem, motorista, placa, transportadora)
        if busca_filter:
            agendamentos = agendamentos.filter(
                Q(ordem__icontains=busca_filter) |
                Q(motorista__nome__icontains=busca_filter) |
                Q(placa_veiculo__icontains=busca_filter) |
                Q(transportadora__nome__icontains=busca_filter)
            )
        
        # Filtro por status
        if status_filter:
            agendamentos = agendamentos.filter(status_geral=status_filter)
        
        # Ordenação padrão
        agendamentos = agendamentos.order_by('-data_agendada', 'horario_agendado')
        
        transportadoras = Transportadora.objects.all().order_by('nome')
        motoristas = Motorista.objects.all().order_by('nome')

        return render(request, 'lista_agendamentos.html', {
            'agendamentos': agendamentos,
            'transportadoras': transportadoras,
            'motoristas': motoristas,
            'data_selecionada': data_filter,
            'filtros_aplicados': {
                'tipo': tipo_filter,
                'transportadora': transportadora_filter,
                'data': data_filter,
                'busca': busca_filter,
                'status': status_filter,
            }
        })
    except Exception as e:
        logger.error(f"Erro na lista_agendamentos: {str(e)}")
        messages.error(request, 'Erro ao carregar lista.')
        data_hoje = timezone_today().strftime('%Y-%m-%d')
        return render(request, 'lista_agendamentos.html', {
            'agendamentos': [], 
            'transportadoras': [],
            'data_selecionada': data_hoje
        })


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('agendamentos')
def exportar_agendamentos(request):
    """
    Exporta os agendamentos filtrados para Excel com informações completas
    Aplica os mesmos filtros da view lista_agendamentos
    Inclui todas as etapas do processo, datas, horários e usuários responsáveis
    """
    try:
        # Query base otimizada com select_related para evitar N+1 queries
        agendamentos = Agendamento.objects.all().select_related(
            'transportadora', 
            'motorista',
            'portaria_liberado_por',
            'checklist_preenchido_por',
            'armazem_confirmado_por',
            'onda_liberado_por',
            'documentos_liberado_por',
            'criado_por',
            'portaria_chegada_armazem_por',
            'armazem_saida_por'
        )
        
        # Aplicar filtros baseados nos parâmetros GET (mesmos da lista)
        tipo_filter = request.GET.get('tipo')
        transportadora_filter = request.GET.get('transportadora')
        data_filter = request.GET.get('data')
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')
        busca_filter = request.GET.get('busca')
        status_filter = request.GET.get('status')
        
        filename_data = ""
        
        # Lógica de filtro de data (Range > Single > Hoje)
        if data_inicio and data_fim:
            try:
                dt_ini = datetime.strptime(data_inicio, '%Y-%m-%d').date()
                dt_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
                agendamentos = agendamentos.filter(data_agendada__range=[dt_ini, dt_fim])
                filename_data = f"{data_inicio.replace('-','')}_{data_fim.replace('-','')}"
            except ValueError:
                data_obj = timezone_today()
                agendamentos = agendamentos.filter(data_agendada=data_obj)
                filename_data = data_obj.strftime('%Y%m%d')
        elif data_filter:
            try:
                data_obj = datetime.strptime(data_filter, '%Y-%m-%d').date()
                agendamentos = agendamentos.filter(data_agendada=data_obj)
                filename_data = data_filter.replace('-', '')
            except ValueError:
                data_obj = timezone_today()
                agendamentos = agendamentos.filter(data_agendada=data_obj)
                filename_data = data_obj.strftime('%Y%m%d')
        else:
            # Se não houver data nem range, usar a data atual como padrão
            data_obj = timezone_today()
            agendamentos = agendamentos.filter(data_agendada=data_obj)
            filename_data = data_obj.strftime('%Y%m%d')
        
        # Filtro por tipo (coleta/entrega)
        if tipo_filter:
            agendamentos = agendamentos.filter(tipo=tipo_filter)
        
        # Filtro por transportadora
        if transportadora_filter:
            agendamentos = agendamentos.filter(transportadora_id=transportadora_filter)
        
        # Filtro por busca (ordem, motorista, placa, transportadora)
        if busca_filter:
            agendamentos = agendamentos.filter(
                Q(ordem__icontains=busca_filter) |
                Q(motorista__nome__icontains=busca_filter) |
                Q(placa_veiculo__icontains=busca_filter) |
                Q(transportadora__nome__icontains=busca_filter)
            )
        
        # Filtro por status
        if status_filter:
            agendamentos = agendamentos.filter(status_geral=status_filter)
        
        # Ordenação padrão
        agendamentos = agendamentos.order_by('-data_agendada', 'horario_agendado')
        
        # Criar workbook Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Agendamentos"
        
        # Função auxiliar para formatar datetime
        def formatar_datetime(dt):
            if dt:
                try:
                    return django_timezone.localtime(dt).strftime('%d/%m/%Y %H:%M')
                except Exception:
                    # Fallback caso não seja aware ou ocorra erro
                    return dt.strftime('%d/%m/%Y %H:%M')
            return ''
        
        def formatar_time(t):
            if t:
                return t.strftime('%H:%M')
            return ''
        
        def formatar_date(d):
            if d:
                return d.strftime('%d/%m/%Y')
            return ''
        
        def get_user_name(user):
            if user:
                return user.get_full_name() or user.username
            return ''
        
        # Cabeçalhos completos
        ws.append([
            'Ordem',
            'Motorista',
            'Telefone Motorista',
            'Data Agendada',
            'Horário Agendado',
            'Tipo',
            'Placa do Veículo',
            'Tipo de Veículo',
            'Transportadora',
            'CNPJ Transportadora',
            'Telefone Transportadora',
            'Peso (kg)',
            'Status Geral',
            'Status Onda',
            # 1. Portaria
            'Portaria - Liberação',
            'Portaria - Liberado Por',
            # 2. Portaria Chegada Armazém
            'Portaria - Chegada Armazém',
            'Portaria - Chegada Armazém Por',
            # 3. Checklist
            'Checklist - Data/Hora',
            'Checklist - Preenchido Por',
            'Checklist - Número',
            'Checklist - Observação',
            # 4. Onda
            'Onda - Data/Hora',
            'Onda - Liberado Por',
            # 5. Início Armazém (Chegada)
            'Armazém - Início',
            'Armazém - Início Por',
            # 6. Fim Armazém (Saída)
            'Armazém - Fim',
            'Armazém - Fim Por',
            'Armazém - Fim Observação',
            # 7. Documentos
            'Documentos - Liberação',
            'Documentos - Liberado Por',
            'Documentos - Observação',
            # Auditoria
            'Criado Em',
            'Criado Por',
            'Atualizado Em',
            # Observações
            'Observações Gerais',
            'Coluna AD'
        ])
        
        # Dados completos
        for agendamento in agendamentos:
            horario_str = 'ENCAIXE' if agendamento.horario_agendado.strftime('%H:%M') == '00:00' else agendamento.horario_agendado.strftime('%H:%M')
            
            ws.append([
                agendamento.ordem,
                agendamento.motorista.nome,
                agendamento.motorista.telefone or '',
                formatar_date(agendamento.data_agendada),
                horario_str,
                agendamento.get_tipo_display(),
                agendamento.placa_veiculo,
                agendamento.get_tipo_veiculo_display(),
                agendamento.transportadora.nome,
                agendamento.transportadora.cnpj or '',
                agendamento.transportadora.telefone or '',
                float(agendamento.peso),
                agendamento.get_status_geral_display(),
                agendamento.get_onda_status_display(),
                # 1. Portaria
                formatar_datetime(agendamento.portaria_liberacao),
                get_user_name(agendamento.portaria_liberado_por),
                # 2. Portaria Chegada Armazém
                formatar_datetime(agendamento.portaria_chegada_armazem),
                get_user_name(agendamento.portaria_chegada_armazem_por),
                # 3. Checklist
                formatar_datetime(agendamento.checklist_data),
                get_user_name(agendamento.checklist_preenchido_por),
                agendamento.checklist_numero or '',
                agendamento.checklist_observacao or '',
                # 4. Onda
                formatar_datetime(agendamento.onda_liberacao),
                get_user_name(agendamento.onda_liberado_por),
                # 5. Armazém Início
                formatar_datetime(agendamento.armazem_chegada),
                get_user_name(agendamento.armazem_confirmado_por),
                # 6. Armazém Fim
                formatar_datetime(agendamento.armazem_saida),
                get_user_name(agendamento.armazem_saida_por),
                agendamento.armazem_saida_observacao or '',
                # 7. Documentos
                formatar_datetime(agendamento.documentos_liberacao),
                get_user_name(agendamento.documentos_liberado_por),
                agendamento.documentos_observacao or '',
                # Auditoria
                formatar_datetime(agendamento.criado_em),
                get_user_name(agendamento.criado_por),
                formatar_datetime(agendamento.atualizado_em),
                # Observações
                agendamento.observacoes or '',
                agendamento.coluna_ad or ''
            ])
        
        # Ajustar largura das colunas automaticamente
        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
            max_length = 0
            column = get_column_letter(idx)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limitar a 50 caracteres
            ws.column_dimensions[column].width = adjusted_width
        
        # Preparar resposta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nome do arquivo com data
        response['Content-Disposition'] = f'attachment; filename=agendamentos_completo_{filename_data}.xlsx'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Erro ao exportar agendamentos: {str(e)}")
        messages.error(request, f'Erro ao exportar agendamentos: {str(e)}')
        return redirect('rondonopolis:lista_agendamentos')


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
def configuracoes_perfil(request):
    from .models import PreferenciaNotificacaoUsuario, ConfiguracaoNotificacao
    
    # Buscar ou criar preferências do usuário
    try:
        preferencias = PreferenciaNotificacaoUsuario.objects.get(usuario=request.user)
    except PreferenciaNotificacaoUsuario.DoesNotExist:
        preferencias = PreferenciaNotificacaoUsuario.objects.create(
            usuario=request.user,
            receber_email=True,
            receber_whatsapp=True,
            receber_navegador=True
        )
    
    # Buscar configuração de notificação (email/whatsapp configurados pelo admin)
    try:
        config = ConfiguracaoNotificacao.objects.get(usuario=request.user)
    except ConfiguracaoNotificacao.DoesNotExist:
        config = None
    
    # Se for POST, salvar apenas preferências (usuário não pode alterar email/telefone)
    if request.method == 'POST':
        receber_email = request.POST.get('receber_email') == 'on'
        receber_whatsapp = request.POST.get('receber_whatsapp') == 'on'
        receber_navegador = request.POST.get('receber_navegador') == 'on'
        
        preferencias.receber_email = receber_email
        preferencias.receber_whatsapp = receber_whatsapp
        preferencias.receber_navegador = receber_navegador
        preferencias.save()
        
        messages.success(request, 'Preferências de notificação salvas com sucesso!')
        return redirect('rondonopolis:configuracoes_perfil')
    
    return render(request, 'configeperfil.html', {
        'preferencias': preferencias,
        'config_notificacao': config,
    })


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
def importar_agendamentos_view(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Use POST'})

    arquivo = request.FILES.get('arquivo')
    if not arquivo:
        return JsonResponse({'success': False, 'error': 'Nenhum arquivo enviado'})

    if not arquivo.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({'success': False, 'error': 'Apenas arquivos Excel'})

    if arquivo.size > 10 * 1024 * 1024:
        return JsonResponse({'success': False, 'error': 'Arquivo muito grande (máx 10MB)'})

    try:
        resultado = importar_agendamentos(arquivo)
        return JsonResponse(resultado)
    except Exception as e:
        logger.error(f"Erro na importação: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erro: {str(e)}'})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
def baixar_modelo_importacao(request):
    colunas = ['ORDEM', 'MOTORISTA', 'DATA AGENDAMENTO', 'TIPO', 'TRANSPORTADORA', 'PLACA', 'VEICULO', 'PESO']
    exemplo = [{
        'ORDEM': 'IMP001', 'MOTORISTA': 'João Silva', 'DATA AGENDAMENTO': '20/12/2025',
        'TIPO': 'entrega', 'TRANSPORTADORA': 'Trans ABC', 'PLACA': 'ABC1D23',
        'VEICULO': 'carreta', 'PESO': '28000'
    }]

    df = pd.DataFrame(exemplo, columns=colunas)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Modelo')
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modelo_importacao.xlsx"'
    return response


@require_http_methods(["GET"])
def documento_impressao(request):
    """
    Gera HTML do documento de internamento (agora sem restrição de saída)
    """
    agendamento_id = request.GET.get('id')
    if not agendamento_id:
        return JsonResponse({'success': False, 'error': 'ID não informado'})

    try:
        agendamento = get_object_or_404(Agendamento, id=agendamento_id)
        # Aplicar a mesma lógica de ordenação das etapas do template de acompanhamento
        etapas_ordenadas = get_etapas_ordenadas(agendamento)
        html = render_to_string('documento_internamento.html', {
            'agendamento': agendamento,
            'etapas_ordenadas': etapas_ordenadas
        })
        return JsonResponse({'success': True, 'html': html})
    except Exception as e:
        logger.error(f"Erro no documento_impressao: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Erro ao gerar documento'})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@require_http_methods(["GET"])
def dados_etapas_agendamento(request, agendamento_id):
    """
    Retorna os dados das etapas de um agendamento para preencher o formulário de gerenciamento
    Os horários são salvos diretamente em Rondonópolis no banco
    """
    try:
        from .utils import converter_para_timezone_rdn
        
        agendamento = get_object_or_404(Agendamento, id=agendamento_id)
        
        # Os horários são salvos diretamente em Rondonópolis (naive)
        # Converter para aware em Rondonópolis para formatar corretamente
        portaria_liberacao = converter_para_timezone_rdn(agendamento.portaria_liberacao) if agendamento.portaria_liberacao else None
        portaria_chegada_armazem = converter_para_timezone_rdn(agendamento.portaria_chegada_armazem) if agendamento.portaria_chegada_armazem else None
        checklist_data = converter_para_timezone_rdn(agendamento.checklist_data) if agendamento.checklist_data else None
        armazem_chegada = converter_para_timezone_rdn(agendamento.armazem_chegada) if agendamento.armazem_chegada else None
        armazem_saida = converter_para_timezone_rdn(agendamento.armazem_saida) if agendamento.armazem_saida else None
        onda_liberacao = converter_para_timezone_rdn(agendamento.onda_liberacao) if agendamento.onda_liberacao else None
        documentos_liberacao = converter_para_timezone_rdn(agendamento.documentos_liberacao) if agendamento.documentos_liberacao else None
        
        # Preparar dados para retornar (formato ISO para datetime-local)
        dados = {
            'id': agendamento.id,
            'tipo': agendamento.tipo,  # Adicionado tipo para controle no frontend
            'portaria_liberacao': portaria_liberacao.strftime('%Y-%m-%dT%H:%M') if portaria_liberacao else None,
            'portaria_chegada_armazem': portaria_chegada_armazem.strftime('%Y-%m-%dT%H:%M') if portaria_chegada_armazem else None,
            'checklist_numero': agendamento.checklist_numero or '',
            'checklist_data': checklist_data.strftime('%Y-%m-%dT%H:%M') if checklist_data else None,
            'checklist_observacao': agendamento.checklist_observacao or '',
            'armazem_chegada': armazem_chegada.strftime('%Y-%m-%dT%H:%M') if armazem_chegada else None,
            'armazem_saida': armazem_saida.strftime('%Y-%m-%dT%H:%M') if armazem_saida else None,
            'armazem_saida_observacao': agendamento.armazem_saida_observacao or '',
            'onda_status': agendamento.onda_status,
            'onda_liberacao': onda_liberacao.strftime('%Y-%m-%dT%H:%M') if onda_liberacao else None,
            'documentos_liberacao': documentos_liberacao.strftime('%Y-%m-%dT%H:%M') if documentos_liberacao else None,
            'documentos_observacao': agendamento.documentos_observacao or '',
        }
        
        return JsonResponse({'success': True, 'agendamento': dados})
    except Exception as e:
        logger.error(f"Erro ao buscar dados das etapas: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Erro ao buscar dados'})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@require_http_methods(["POST"])
def salvar_etapas_agendamento(request):
    """
    Salva as alterações nas etapas de um agendamento
    """
    try:
        agendamento_id = request.POST.get('agendamento_id')
        if not agendamento_id:
            return JsonResponse({'success': False, 'error': 'ID do agendamento não informado'})
        
        agendamento = get_object_or_404(Agendamento, id=agendamento_id)
        
        # Converter strings de data/hora para objetos datetime
        # Os valores vêm do datetime-local no formato YYYY-MM-DDTHH:MM no fuso de Rondonópolis
        from datetime import datetime
        import pytz
        from .utils import TIMEZONE_RONDONOPOLIS
        
        # Portaria
        portaria_liberacao_str = request.POST.get('portaria_liberacao', '').strip()
        etapa_rapida = request.POST.get('etapa_rapida', '').strip()
        
        # Se for confirmação rápida e não tiver data, usar horário atual
        if etapa_rapida == 'portaria' and not portaria_liberacao_str:
            portaria_liberacao_str = timezone_now().strftime('%Y-%m-%dT%H:%M')
        
        if portaria_liberacao_str:
            try:
                # O valor vem do datetime-local no formato YYYY-MM-DDTHH:MM
                # O usuário está preenchendo o horário de Rondonópolis
                # Salvar diretamente como naive no horário de Rondonópolis
                dt_naive = datetime.strptime(portaria_liberacao_str, '%Y-%m-%dT%H:%M')
                agendamento.portaria_liberacao = dt_naive
                agendamento.portaria_liberado_por = request.user
            except Exception as e:
                logger.error(f"Erro ao converter portaria_liberacao: {str(e)}")
        
        # Checklist
        agendamento.checklist_numero = request.POST.get('checklist_numero', '').strip()
        
        checklist_data_str = request.POST.get('checklist_data', '').strip()
        
        # Se for confirmação rápida e não tiver data, usar horário atual
        if etapa_rapida == 'checklist' and not checklist_data_str:
            checklist_data_str = timezone_now().strftime('%Y-%m-%dT%H:%M')
        
        if checklist_data_str:
            try:
                # O usuário está preenchendo o horário de Rondonópolis
                # Salvar diretamente como naive no horário de Rondonópolis
                dt_naive = datetime.strptime(checklist_data_str, '%Y-%m-%dT%H:%M')
                agendamento.checklist_data = dt_naive
                agendamento.checklist_preenchido_por = request.user
            except Exception as e:
                logger.error(f"Erro ao converter checklist_data: {str(e)}")
        
        # Observação do checklist
        agendamento.checklist_observacao = request.POST.get('checklist_observacao', '').strip()
        
        # Armazém
        armazem_chegada_str = request.POST.get('armazem_chegada', '').strip()
        
        # Se for confirmação rápida e não tiver data, usar horário atual
        if etapa_rapida == 'armazem' and not armazem_chegada_str:
            armazem_chegada_str = timezone_now().strftime('%Y-%m-%dT%H:%M')
        
        if armazem_chegada_str:
            try:
                # O usuário está preenchendo o horário de Rondonópolis
                # Salvar diretamente como naive no horário de Rondonópolis
                dt_naive = datetime.strptime(armazem_chegada_str, '%Y-%m-%dT%H:%M')
                agendamento.armazem_chegada = dt_naive
                agendamento.armazem_confirmado_por = request.user
            except Exception as e:
                logger.error(f"Erro ao converter armazem_chegada: {str(e)}")
        
        # Armazém Saída
        armazem_saida_str = request.POST.get('armazem_saida', '').strip()
        
        if armazem_saida_str:
            try:
                dt_naive = datetime.strptime(armazem_saida_str, '%Y-%m-%dT%H:%M')
                agendamento.armazem_saida = dt_naive
                agendamento.armazem_saida_por = request.user
            except Exception as e:
                logger.error(f"Erro ao converter armazem_saida: {str(e)}")
        
        # Observação do armazém (saída)
        agendamento.armazem_saida_observacao = request.POST.get('armazem_observacao', '').strip()
        
        # Onda
        agendamento.onda_status = request.POST.get('onda_status', 'aguardando')
        
        onda_liberacao_str = request.POST.get('onda_liberacao', '').strip()
        
        # Se for confirmação rápida e não tiver data, usar horário atual
        if etapa_rapida == 'onda' and not onda_liberacao_str:
            onda_liberacao_str = timezone_now().strftime('%Y-%m-%dT%H:%M')
            agendamento.onda_status = 'liberado'
        
        if onda_liberacao_str:
            try:
                # O usuário está preenchendo o horário de Rondonópolis
                # Salvar diretamente como naive no horário de Rondonópolis
                dt_naive = datetime.strptime(onda_liberacao_str, '%Y-%m-%dT%H:%M')
                agendamento.onda_liberacao = dt_naive
                agendamento.onda_liberado_por = request.user
                if etapa_rapida == 'onda':
                    agendamento.onda_status = 'liberado'
            except Exception as e:
                logger.error(f"Erro ao converter onda_liberacao: {str(e)}")
        
        # Documentos
        documentos_liberacao_str = request.POST.get('documentos_liberacao', '').strip()
        
        # Se for confirmação rápida e não tiver data, usar horário atual
        if etapa_rapida == 'documentos' and not documentos_liberacao_str:
            documentos_liberacao_str = timezone_now().strftime('%Y-%m-%dT%H:%M')
        
        if documentos_liberacao_str:
            try:
                # O usuário está preenchendo o horário de Rondonópolis
                # Salvar diretamente como naive no horário de Rondonópolis
                dt_naive = datetime.strptime(documentos_liberacao_str, '%Y-%m-%dT%H:%M')
                agendamento.documentos_liberacao = dt_naive
                agendamento.documentos_liberado_por = request.user
            except Exception as e:
                logger.error(f"Erro ao converter documentos_liberacao: {str(e)}")
        
        # Observação dos documentos
        agendamento.documentos_observacao = request.POST.get('documentos_observacao', '').strip()
        
        # Atualizar status geral
        agendamento.atualizar_status_geral(request.user)
        agendamento.save()
        
        return JsonResponse({'success': True, 'message': 'Etapas atualizadas com sucesso!'})
        
    except Exception as e:
        logger.error(f"Erro ao salvar etapas: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erro ao salvar: {str(e)}'})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
def motoristas_telefone(request):
    """
    View para cadastro/edição de telefones dos motoristas
    """
    motoristas = Motorista.objects.all().order_by('nome')
    
    if request.method == 'POST':
        motorista_id = request.POST.get('motorista_id')
        telefone = request.POST.get('telefone', '').strip()
        
        try:
            motorista = Motorista.objects.get(id=motorista_id)
            motorista.telefone = telefone
            motorista.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Telefone de {motorista.nome} atualizado com sucesso!'
            })
        except Motorista.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Motorista não encontrado'
            })
    
    return render(request, 'motoristas_telefone.html', {
        'motoristas': motoristas
    })


@login_required
@require_http_methods(["POST"])
def chamar_motorista(request):
    """
    View para chamar motorista via WhatsApp
    Também registra a chamada na fila de áudio (mesmo sem telefone)
    """
    from django.core.cache import cache
    import uuid
    
    agendamento_id = request.POST.get('agendamento_id')
    
    if not agendamento_id:
        return JsonResponse({'success': False, 'error': 'ID do agendamento não fornecido'})
    
    try:
        agendamento = get_object_or_404(Agendamento, id=agendamento_id)
        motorista = agendamento.motorista
        
        # Preparar dados da chamada para a fila de áudio (sempre adiciona, mesmo sem telefone)
        timestamp_unico = timezone_now().isoformat()
        chamada_audio = {
            'id': str(uuid.uuid4()),
            'agendamento_id': agendamento.id,
            'motorista_nome': motorista.nome,
            'placa': agendamento.placa_veiculo,
            'tipo': agendamento.get_tipo_display(),
            'transportadora': agendamento.transportadora.nome,
            'timestamp': timestamp_unico,
        }
        
        # Adicionar à fila de chamadas de áudio
        fila_chamadas = cache.get('fila_chamadas_motorista', [])
        if not isinstance(fila_chamadas, list):
            fila_chamadas = []
        fila_chamadas.append(chamada_audio)
        cache.set('fila_chamadas_motorista', fila_chamadas, timeout=300)  # 5 minutos
        
        # Se tiver telefone, tentar enviar WhatsApp
        if motorista.telefone:
            # Limpar e formatar telefone
            telefone = motorista.telefone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
            
            # Formatar data e hora
            data_formatada = agendamento.data_agendada.strftime('%d/%m/%Y')
            hora_formatada = agendamento.horario_agendado.strftime('%H:%M')
            data_hora = f"{data_formatada} às {hora_formatada}"
            
            # Obter tipo do serviço
            tipo_servico = agendamento.get_tipo_display()
            
            # Mensagem padrão
            mensagem = (
                f"Olá, {motorista.nome}!\n\n"
                f"Por gentileza, dirija-se à portaria com o veículo para atendimento conforme agendamento:\n\n"
                f"• Placa: {agendamento.placa_veiculo}\n"
                f"• Data e horário: {data_hora}\n"
                f"• Serviço: {tipo_servico}\n\n"
                f"Obrigado!\n\n"
                f"TLOGpainel\n"
                f"Transcamila Cargas e Armazéns Gerais LTDA"
            )
            
            # Tentar enviar via API primeiro
            api_result = enviar_whatsapp_api(telefone, mensagem)
            
            if api_result.get('success'):
                # Mensagem enviada com sucesso via API
                return JsonResponse({
                    'success': True,
                    'message': f'Mensagem enviada com sucesso para {motorista.nome}',
                    'sent_via_api': True
                })
            else:
                # Se a API não estiver configurada ou falhar, usar método alternativo (link WhatsApp Web)
                logger.warning(f"API WhatsApp não disponível, usando método alternativo: {api_result.get('error')}")
                
                # Codificar mensagem para URL
                mensagem_encoded = requests.utils.quote(mensagem)
                
                # URL do WhatsApp Web como fallback
                whatsapp_url = f"https://wa.me/55{telefone}?text={mensagem_encoded}"
                
                return JsonResponse({
                    'success': True,
                    'whatsapp_url': whatsapp_url,
                    'message': f'Abrindo WhatsApp para {motorista.nome}...',
                    'sent_via_api': False,
                    'api_warning': api_result.get('error', 'API não configurada')
                })
        else:
            # Sem telefone, mas chamada adicionada à fila de áudio
            return JsonResponse({
                'success': True,
                'message': f'Chamada de áudio registrada para {motorista.nome} (sem telefone cadastrado)',
                'sent_via_api': False,
                'sem_telefone': True
            })
        
    except Exception as e:
        logger.error(f"Erro em chamar_motorista: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Erro interno do servidor'})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
def tela_chamada_motorista(request):
    """
    Página de chamada de motoristas - tela separada em tempo real
    Fica aberta e verifica novas chamadas periodicamente
    """
    return render(request, 'tela_chamada_motorista.html', {
        'mostrar_chamada': False,
    })


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@require_http_methods(["GET"])
def verificar_chamada_motorista(request):
    """
    Endpoint para verificar se há nova chamada de motorista na fila
    Retorna a próxima chamada da fila se houver
    """
    from django.core.cache import cache
    
    # Obter fila de chamadas
    fila_chamadas = cache.get('fila_chamadas_motorista', [])
    if not isinstance(fila_chamadas, list):
        fila_chamadas = []
    
    # Verificar se há chamadas processadas (para evitar reprocessar)
    chamadas_processadas = cache.get('chamadas_processadas_motorista', set())
    if not isinstance(chamadas_processadas, set):
        chamadas_processadas = set()
    
    # Procurar próxima chamada não processada
    proxima_chamada = None
    for chamada in fila_chamadas:
        if chamada.get('id') not in chamadas_processadas:
            proxima_chamada = chamada
            break
    
    if proxima_chamada:
        # Marcar como processada
        chamadas_processadas.add(proxima_chamada.get('id'))
        cache.set('chamadas_processadas_motorista', chamadas_processadas, timeout=300)
        
        # Retornar os dados da chamada
        return JsonResponse({
            'success': True,
            'tem_chamada': True,
            'chamada': proxima_chamada
        })
    else:
        return JsonResponse({
            'success': True,
            'tem_chamada': False
        })


@login_required
@require_http_methods(["POST"])
def iniciar_chamada_motorista(request):
    """
    View para iniciar a chamada de um motorista
    Retorna informações do motorista para exibir na tela de chamada
    """
    # Verificar se o usuário é administrador
    if not request.user.is_superuser and not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Acesso restrito apenas para administradores.'})
    
    motorista_id = request.POST.get('motorista_id')
    agendamento_id = request.POST.get('agendamento_id')
    
    if not motorista_id and not agendamento_id:
        return JsonResponse({'success': False, 'error': 'ID do motorista ou agendamento não fornecido'})
    
    try:
        if agendamento_id:
            agendamento = get_object_or_404(Agendamento, id=agendamento_id)
            motorista = agendamento.motorista
            agendamento_info = {
                'id': agendamento.id,
                'placa': agendamento.placa_veiculo,
                'tipo': agendamento.get_tipo_display(),
                'transportadora': agendamento.transportadora.nome,
                'horario': str(agendamento.horario_agendado),
            }
        else:
            motorista = get_object_or_404(Motorista, id=motorista_id)
            # Buscar o primeiro agendamento do dia para este motorista
            agendamento = Agendamento.objects.filter(
                motorista=motorista,
                data_agendada=timezone_now().date()
            ).first()
            agendamento_info = {
                'id': agendamento.id if agendamento else None,
                'placa': agendamento.placa_veiculo if agendamento else 'N/A',
                'tipo': agendamento.get_tipo_display() if agendamento else 'N/A',
                'transportadora': agendamento.transportadora.nome if agendamento else 'N/A',
                'horario': str(agendamento.horario_agendado) if agendamento else 'N/A',
            }
        
        return JsonResponse({
            'success': True,
            'motorista': {
                'id': motorista.id,
                'nome': motorista.nome,
                'telefone': motorista.telefone or '',
            },
            'agendamento': agendamento_info,
        })
        
    except Exception as e:
        logger.error(f"Erro em iniciar_chamada_motorista: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Erro interno do servidor'})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')



@login_required
@require_http_methods(["POST"])
def preencher_checklist(request):
    """
    Preenche o CheckList (apenas o número é obrigatório)
    Aceita ação individual ou em massa
    """
    # Suporta seleção múltipla ou ação individual
    ids = request.POST.getlist('agendamentos_ids[]')
    if not ids:
        single_id = request.POST.get('agendamento_id')
        if single_id:
            ids = [single_id]

    numero_checklist = request.POST.get('numero_checklist', '').strip()
    observacao = request.POST.get('observacao', '').strip()

    if not ids:
        return JsonResponse({'success': False, 'error': 'Nenhum veículo selecionado'})

    if not numero_checklist:
        return JsonResponse({'success': False, 'error': 'Número do CheckList é obrigatório'})

    sucessos = 0
    erros = 0

    # IMPORTANTE: select_for_update() precisa estar dentro de uma transação no MySQL
    from django.db import transaction
    
    for agendamento_id in ids:
        try:
            # Usar transaction.atomic() para garantir que select_for_update() funcione no MySQL
            with transaction.atomic():
                ag = Agendamento.objects.select_for_update().get(id=agendamento_id)

                # Evita preenchimento duplicado
                if ag.checklist_data is not None:
                    erros += 1
                    continue

                ag.checklist_numero = numero_checklist
                # Usar timezone_now() que retorna datetime naive no fuso de Rondonópolis
                # Salvar diretamente no horário de Rondonópolis (sem conversão para UTC)
                agora_rdn = timezone_now()
                ag.checklist_data = agora_rdn
                ag.checklist_preenchido_por = request.user
                ag.checklist_observacao = observacao if observacao else None
                # atualizar_status_geral() já faz save() internamente, mas precisamos salvar todos os campos
                ag.atualizar_status_geral()
                # Salvar novamente para garantir que todos os campos sejam persistidos
                ag.save()

            # Enviar notificações fora da transação (assíncrono, não bloqueia resposta)
            try:
                enviar_notificacao_etapa(ag, 'checklist', request.user)
            except Exception as e:
                logger.error(f"Erro ao enviar notificação: {str(e)}")

            sucessos += 1
            logger.info(f"CHECKLIST #{numero_checklist} - {request.user} - Placa: {ag.placa_veiculo}")

        except Agendamento.DoesNotExist:
            erros += 1
        except Exception as e:
            logger.error(f"Erro checklist ID {agendamento_id}: {e}")
            import traceback
            logger.error(traceback.format_exc())
            erros += 1

    if sucessos > 0:
        return JsonResponse({
            'success': True,
            'message': f'CheckList #{numero_checklist} registrado em {sucessos} veículo(s)!'
        })
    else:
        return JsonResponse({
            'success': False,
            'error': 'Nenhum CheckList foi registrado.'
        })





@login_required
@require_http_methods(["POST"])
def armazem_registrar_entrada(request):
    """
    Registra a chegada do veículo no armazém (após CheckList)
    Usa os campos já existentes: armazem_chegada e armazem_confirmado_por
    """
    ids = request.POST.getlist('ids[]')
    if not ids:
        return JsonResponse({'success': False, 'error': 'Nenhum veículo selecionado'})

    sucessos = 0
    for ag_id in ids:
        try:
            with transaction.atomic():
                ag = Agendamento.objects.select_for_update().get(id=ag_id)

                # Já entrou? Pula
                if ag.armazem_chegada is not None:
                    continue

                # Registra entrada
                # Usar timezone_now() que retorna datetime naive no fuso de Rondonópolis
                # Salvar diretamente no horário de Rondonópolis (sem conversão para UTC)
                agora_rdn = timezone_now()
                ag.armazem_chegada = agora_rdn
                ag.armazem_confirmado_por = request.user
                ag.atualizar_status_geral()
                ag.save()

                # Enviar notificações (assíncrono, não bloqueia resposta)
                # SOLICITAÇÃO 12/01: Desativado para evitar envio na entrada do armazém (duplicata de função)
                # enviar_notificacao_etapa(ag, 'armazem', request.user)

                sucessos += 1
                logger.info(f"ARMAZÉM - Entrada registrada por {request.user} - Placa: {ag.placa_veiculo}")

        except Agendamento.DoesNotExist:
            continue
        except Exception as e:
            logger.error(f"Erro ao registrar entrada no armazém (ID {ag_id}): {e}")

    if sucessos > 0:
        return JsonResponse({
            'success': True,
            'message': f'Entrada no armazém registrada em {sucessos} veículo(s)!'
        })
    else:
        return JsonResponse({
            'success': False,
            'error': 'Nenhum veículo foi registrado (já estavam no armazém ou erro)'
        })


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@require_http_methods(["GET"])
def exibir_processo_detalhes(request, agendamento_id):
    """
    Retorna todos os detalhes de um processo para exibição em modal
    """
    try:
        agendamento = Agendamento.objects.select_related(
            'transportadora', 'motorista', 
            'portaria_liberado_por', 'checklist_preenchido_por',
            'armazem_confirmado_por', 'armazem_saida_por',
            'onda_liberado_por', 'documentos_liberado_por'
        ).get(id=agendamento_id)
        
        # Obter etapas ordenadas
        etapas = get_etapas_ordenadas(agendamento)
        
        # Preparar dados do processo
        dados = {
            'success': True,
            'processo': {
                'id': agendamento.id,
                'ordem': agendamento.ordem,
                'tipo': agendamento.tipo,
                'tipo_display': agendamento.get_tipo_display(),
                'motorista': agendamento.motorista.nome if agendamento.motorista else '',
                'placa_veiculo': agendamento.placa_veiculo,
                'tipo_veiculo': agendamento.get_tipo_veiculo_display(),
                'transportadora': agendamento.transportadora.nome if agendamento.transportadora else '',
                'data_agendada': agendamento.data_agendada.strftime('%d/%m/%Y') if agendamento.data_agendada else '',
                'horario_agendado': agendamento.horario_agendado.strftime('%H:%M') if agendamento.horario_agendado else '',
                'peso': agendamento.peso or 0,
                'status_geral': agendamento.status_geral,
                'status_geral_display': agendamento.get_status_geral_display(),
                
                # Observações do checklist
                'checklist_numero': agendamento.checklist_numero or '',
                'checklist_observacao': agendamento.checklist_observacao or '',
                
                # Observações do armazém
                'armazem_observacao': agendamento.armazem_saida_observacao or '',
                
                # Observações dos documentos
                'documentos_observacao': agendamento.documentos_observacao or '',

                # Chegada no Armazém (Portaria)
                'portaria_chegada_armazem': agendamento.portaria_chegada_armazem.strftime('%d/%m/%Y %H:%M') if agendamento.portaria_chegada_armazem else None,
                'portaria_chegada_armazem_por': agendamento.portaria_chegada_armazem_por.get_full_name() if agendamento.portaria_chegada_armazem_por else None,
                
                # Etapas
                'etapas': [
                    {
                        'nome': etapa['nome'],
                        'concluida': etapa['concluida'],
                        'disponivel': etapa['disponivel'],
                        'tipo': etapa.get('tipo', 'sequencial'),
                        'data': etapa['conclusao_formatada'] if etapa['concluida'] else None,
                        'usuario': etapa.get('usuario', '') if etapa['concluida'] else '',
                    }
                    for etapa in etapas
                ]
            }
        }
        
        return JsonResponse(dados)
        
    except Agendamento.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Processo não encontrado'
        }, status=404)
    except Exception as e:
        logger.error(f"Erro ao buscar detalhes do processo {agendamento_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': 'Erro ao carregar detalhes do processo'
        }, status=500)


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('processos')
def visualizacao_processos(request):
    """
    View para visualização completa de todos os processos
    """
    try:
        # Obter todos os agendamentos inicialmente
        agendamentos_list = Agendamento.objects.select_related('transportadora', 'motorista').all()
        
        # Aplicar filtros baseados nos parâmetros GET
        status_filter = request.GET.get('status')
        tipo_filter = request.GET.get('tipo')
        data_filter = request.GET.get('data')
        busca_filter = request.GET.get('busca')
        motorista_filter = request.GET.get('motorista')
        
        # Data padrão: hoje (se não houver filtro de data)
        data_hoje = timezone_now().date()
        data_selecionada = data_hoje
        
        # Filtro por status
        if status_filter:
            agendamentos_list = agendamentos_list.filter(status_geral=status_filter)
        
        # Filtro por tipo (coleta/entrega)
        if tipo_filter:
            agendamentos_list = agendamentos_list.filter(tipo=tipo_filter)
        
        # Filtro por motorista
        if motorista_filter:
            agendamentos_list = agendamentos_list.filter(motorista_id=motorista_filter)
        
        # Filtro por data - sempre aplica (padrão é hoje se não especificado)
        if data_filter:
            try:
                data_obj = datetime.strptime(data_filter, '%Y-%m-%d').date()
                data_selecionada = data_obj
                agendamentos_list = agendamentos_list.filter(data_agendada=data_obj)
            except ValueError:
                # Se a data for inválida, usa a data de hoje
                agendamentos_list = agendamentos_list.filter(data_agendada=data_hoje)
        else:
            # Se não houver filtro de data, usa a data de hoje por padrão
            agendamentos_list = agendamentos_list.filter(data_agendada=data_hoje)
        
        # Filtro por busca (ordem, motorista, placa, transportadora)
        if busca_filter:
            agendamentos_list = agendamentos_list.filter(
                Q(ordem__icontains=busca_filter) |
                Q(motorista__nome__icontains=busca_filter) |
                Q(placa_veiculo__icontains=busca_filter) |
                Q(transportadora__nome__icontains=busca_filter)
            )
        
        # Ordenação padrão
        agendamentos_list = agendamentos_list.order_by('-data_agendada', 'horario_agendado')
        
        # Processar cada agendamento para adicionar etapas ordenadas
        agendamentos_com_etapas = []
        for agendamento in agendamentos_list:
            etapas_ordenadas = get_etapas_ordenadas(agendamento)
            
            # Calcular progresso: contar etapas concluídas em sequência
            # Para na primeira etapa não concluída
            etapas_concluidas = 0
            for etapa in etapas_ordenadas:
                if etapa.get('concluida', False):
                    etapas_concluidas += 1
                else:
                    break
            
            total_etapas = len(etapas_ordenadas)
            
            agendamento_dict = {
                'obj': agendamento,
                'etapas_ordenadas': etapas_ordenadas,
                'etapas_concluidas': etapas_concluidas,
                'total_etapas': total_etapas
            }
            agendamentos_com_etapas.append(agendamento_dict)
        
        # Paginação - 10 itens por página
        paginator = Paginator(agendamentos_com_etapas, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Buscar motoristas únicos que têm agendamentos para popular o filtro
        # Usar a mesma query base para garantir que apenas motoristas com agendamentos sejam mostrados
        motoristas_ids = agendamentos_list.values_list('motorista_id', flat=True).distinct()
        motoristas = Motorista.objects.filter(id__in=motoristas_ids).order_by('nome')
        
        # Verificar se o usuário é administrador (grupo)
        is_admin = False
        try:
            grupo_admin = GrupoUsuario.objects.filter(nome='administracao', ativo=True).first()
            if grupo_admin and grupo_admin.usuarios.filter(id=request.user.id).exists():
                is_admin = True
        except Exception:
            pass

        context = {
            'agendamentos': page_obj,
            'data_selecionada': data_selecionada.strftime('%Y-%m-%d'),  # Para preencher o campo de data no template
            'motoristas': motoristas,
            'is_admin': is_admin,
        }
        
        return render(request, 'visualizacao_processos.html', context)
        
    except Exception as e:
        logger.error(f"Erro em visualizacao_processos: {str(e)}")
        messages.error(request, 'Erro ao carregar visualização de processos.')
        data_hoje = timezone_now().date()
        return render(request, 'visualizacao_processos.html', {
            'agendamentos': [],
            'data_selecionada': data_hoje.strftime('%Y-%m-%d')
        })


def get_etapas_ordenadas(agendamento):
    """
    Retorna as etapas ordenadas pela ordem de conclusão
    Converte os datetimes para o fuso horário de Rondonópolis
    """
    from .utils import converter_para_timezone_rdn
    
    etapas = []
    
    # Identificar tipo
    is_entrega = (agendamento.tipo == 'entrega')

    # 1. Portaria - sempre primeira etapa
    portaria_concluida = bool(agendamento.portaria_liberacao)
    portaria_data = converter_para_timezone_rdn(agendamento.portaria_liberacao) if agendamento.portaria_liberacao else None
    portaria_usuario = agendamento.portaria_liberado_por.username if agendamento.portaria_liberado_por else None
    portaria_conclusao = portaria_data.strftime('%d/%m/%Y %H:%M') if portaria_data else None
    etapas.append({
        'nome': 'Portaria',
        'concluida': portaria_concluida,
        'disponivel': True,
        'data': portaria_data,
        'usuario': portaria_usuario,
        'conclusao_formatada': portaria_conclusao,
        'ordem_padrao': 1,
        'tipo': 'sequencial'
    })

    # 2. Chegada no Armazém (Confirmado pela Portaria)
    chegada_concluida = bool(agendamento.portaria_chegada_armazem)
    chegada_disponivel = portaria_concluida
    chegada_data = converter_para_timezone_rdn(agendamento.portaria_chegada_armazem) if agendamento.portaria_chegada_armazem else None
    chegada_usuario = agendamento.portaria_chegada_armazem_por.username if agendamento.portaria_chegada_armazem_por else None
    chegada_conclusao = chegada_data.strftime('%d/%m/%Y %H:%M') if chegada_data else None
    etapas.append({
        'nome': 'Cheg. Arm.',
        'concluida': chegada_concluida,
        'disponivel': chegada_disponivel,
        'data': chegada_data,
        'usuario': chegada_usuario,
        'conclusao_formatada': chegada_conclusao,
        'ordem_padrao': 2,
        'tipo': 'sequencial'
    })

    # 3. Checklist - APENAS PARA COLETA
    checklist_concluida = bool(agendamento.checklist_data)
    if not is_entrega:
        checklist_disponivel = chegada_concluida # Disponível após chegada no armazém
        checklist_data = converter_para_timezone_rdn(agendamento.checklist_data) if agendamento.checklist_data else None
        checklist_usuario = agendamento.checklist_preenchido_por.username if agendamento.checklist_preenchido_por else None
        checklist_conclusao = checklist_data.strftime('%d/%m/%Y %H:%M') if checklist_data else None
        etapas.append({
            'nome': 'Checklist',
            'concluida': checklist_concluida,
            'disponivel': checklist_disponivel,
            'data': checklist_data,
            'usuario': checklist_usuario,
            'conclusao_formatada': checklist_conclusao,
            'ordem_padrao': 3,
            'tipo': 'sequencial'
        })
    
    # 4. Onda (Coleta) ou OD (Entrega)
    # etapa independente
    onda_concluida = bool(agendamento.onda_liberacao)
    onda_data = converter_para_timezone_rdn(agendamento.onda_liberacao) if agendamento.onda_liberacao else None
    
    nome_etapa = 'OD' if is_entrega else 'Onda'
    onda_usuario = agendamento.onda_liberado_por.username if agendamento.onda_liberado_por else None
    onda_conclusao = onda_data.strftime('%d/%m/%Y %H:%M') if onda_data else None
    
    etapas.append({
        'nome': nome_etapa,
        'concluida': onda_concluida,
        'disponivel': True,  # Sempre disponível - pode ser liberada a qualquer momento
        'data': onda_data,
        'usuario': onda_usuario,
        'conclusao_formatada': onda_conclusao,
        'ordem_padrao': 4,
        'tipo': 'independente'
    })
    
    # 5. Início Operação (Entrada Armazém)
    entrada_concluida = bool(agendamento.armazem_chegada)
    
    # Disponibilidade
    if is_entrega:
        entrada_disponivel = chegada_concluida
    else:
        # Coleta: Disponível após Checklist e idealmente após Chegada
        entrada_disponivel = checklist_concluida
        
    entrada_data = converter_para_timezone_rdn(agendamento.armazem_chegada) if agendamento.armazem_chegada else None
    entrada_usuario = agendamento.armazem_confirmado_por.username if agendamento.armazem_confirmado_por else None
    entrada_conclusao = entrada_data.strftime('%d/%m/%Y %H:%M') if entrada_data else None
    etapas.append({
        'nome': 'Início Arm.',
        'concluida': entrada_concluida,
        'disponivel': entrada_disponivel,
        'data': entrada_data,
        'usuario': entrada_usuario,
        'conclusao_formatada': entrada_conclusao,
        'ordem_padrao': 5,
        'tipo': 'sequencial'
    })

    # 6. Fim Operação (Saída Armazém)
    saida_concluida = bool(agendamento.armazem_saida)
    saida_disponivel = entrada_concluida
    
    saida_data = converter_para_timezone_rdn(agendamento.armazem_saida) if agendamento.armazem_saida else None
    saida_usuario = agendamento.armazem_saida_por.username if agendamento.armazem_saida_por else None
    saida_conclusao = saida_data.strftime('%d/%m/%Y %H:%M') if saida_data else None
    etapas.append({
        'nome': 'Fim Arm.',
        'concluida': saida_concluida,
        'disponivel': saida_disponivel,
        'data': saida_data,
        'usuario': saida_usuario,
        'conclusao_formatada': saida_conclusao,
        'ordem_padrao': 6,
        'tipo': 'sequencial'
    })
    
    # 7. Documentos - última etapa
    documentos_concluida = bool(agendamento.documentos_liberacao)
    documentos_disponivel = saida_concluida
    documentos_data = converter_para_timezone_rdn(agendamento.documentos_liberacao) if agendamento.documentos_liberacao else None
    documentos_usuario = agendamento.documentos_liberado_por.username if agendamento.documentos_liberado_por else None
    documentos_conclusao = documentos_data.strftime('%d/%m/%Y %H:%M') if documentos_data else None
    etapas.append({
            'nome': 'Documentos',
        'concluida': documentos_concluida,
        'disponivel': True,  # Sempre disponível - pode ser liberada a qualquer momento
        'data': documentos_data,
        'usuario': documentos_usuario,
        'conclusao_formatada': documentos_conclusao,
        'ordem_padrao': 7,
        'tipo': 'independente'  # Tipo independente pois se move dinamicamente
    })
    
    # Calcular ordem dinâmica para Onda/OD e Documentos baseado nos timestamps
    def calcular_ordem_dinamica(etapa, todas_etapas):
        """
        Calcula a ordem dinâmica baseada no timestamp de conclusão.
        """
        nome = etapa['nome']
        dt_conclusao = etapa.get('data')

        # Se não concluída, define posições padrão de "espera"
        if not etapa['concluida']:
            if nome in ['OD', 'Onda']:
                return 3.5 # Antes do Início Armazém
            if nome == 'Documentos':
                return 7 # Final
            return etapa.get('ordem_padrao', 999)

        # Para Onda/OD - Inserir cronologicamente entre as etapas sequenciais
        if nome in ['OD', 'Onda']:
            if not dt_conclusao:
                return 3.5
                
            # Pegar etapas sequenciais concluídas que têm data
            sequenciais_concluidas = [
                e for e in todas_etapas 
                if e.get('tipo') == 'sequencial' and e.get('concluida') and e.get('data') and e['nome'] not in ['OD', 'Onda', 'Documentos']
            ]
            
            # Encontrar a última etapa sequencial que ocorreu ANTES da Onda
            ordem_anterior_max = 0
            
            for seq in sequenciais_concluidas:
                if seq['data'] <= dt_conclusao:
                    # Esta etapa ocorreu antes ou no mesmo tempo da Onda
                    if seq['ordem_padrao'] > ordem_anterior_max:
                        ordem_anterior_max = seq['ordem_padrao']
            
            if ordem_anterior_max > 0:
                # Fica depois da última etapa sequencial anterior a ela
                return ordem_anterior_max + 0.5
            else:
                # Ocorreu antes de qualquer etapa sequencial concluída
                return 0.5
        
        # Para Documentos - Regras específicas de visibilidade
        elif nome == 'Documentos':
            # Buscar Fim do Armazém
            fim_arm = next((e for e in todas_etapas if e['nome'] == 'Fim Arm.'), None)
            
            if fim_arm and fim_arm.get('concluida') and fim_arm.get('data') and dt_conclusao:
                if dt_conclusao > fim_arm['data']:
                    # Documentos feito DEPOIS do Fim do Armazém -> Final
                    return 7
            
            # Caso contrário (feito antes do fim, ou fim não concluído) -> Entre Início e Fim
            return 5.5
        
        # Para outras etapas, manter ordem padrão
        else:
            return etapa.get('ordem_padrao', 999)
    
    # Calcular ordem dinâmica para cada etapa
    for etapa in etapas:
        etapa['ordem_dinamica'] = calcular_ordem_dinamica(etapa, etapas)
    
    # Ordenar pela ordem dinâmica
    etapas_ordenadas = sorted(
        etapas,
        key=lambda x: x.get('ordem_dinamica', x.get('ordem_padrao', 999))
    )
    
    return etapas_ordenadas


from django.shortcuts import render
from django.db.models import Count, Q, Sum, F
from .models import Agendamento, Transportadora
from datetime import datetime
from .utils import timezone_now, timezone_today

@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('painel')
def processos_painel(request):
    # Verificar se o usuário está no grupo "Monitores"
    is_monitor = False
    try:
        grupo_monitores = GrupoUsuario.objects.filter(nome='monitores', ativo=True).first()
        if grupo_monitores and grupo_monitores.usuarios.filter(id=request.user.id).exists():
            is_monitor = True
    except Exception:
        pass
    
    # Sempre mostrar apenas os agendamentos do dia atual
    data_selecionada = timezone_now().date()
    
    # Query base - apenas agendamentos de hoje
    agendamentos = Agendamento.objects.filter(data_agendada=data_selecionada)
    
    # Estatísticas para os cards
    estatisticas = agendamentos.aggregate(
        coletas=Count('id', filter=Q(tipo='coleta')),
        entregas=Count('id', filter=Q(tipo='entrega')),
        concluidos=Count('id', filter=Q(status_geral='processo_concluido')),
        peso_total=Sum('peso')
    )
    
    # Veículos = soma de coletas + entregas (cada processo é um veículo)
    veiculos_count = (estatisticas['coletas'] or 0) + (estatisticas['entregas'] or 0)
    
    # Preparar agendamentos com etapas - ordenar por atualizado_em desc (mais recente primeiro), depois horario
    agendamentos_com_etapas = []
    for ag in agendamentos.select_related('transportadora', 'motorista', 'portaria_liberado_por', 
                                         'checklist_preenchido_por', 'armazem_confirmado_por',
                                         'onda_liberado_por').order_by(F('atualizado_em').desc(nulls_last=True), 'horario_agendado'):
        etapas = get_etapas_ordenadas(ag)
        
        # Calcular progresso: contar etapas concluídas em sequência
        etapas_concluidas = 0
        for etapa in etapas:
            if etapa.get('concluida', False):
                etapas_concluidas += 1
            else:
                break
        
        total_etapas = len(etapas)
        
        agendamentos_com_etapas.append({
            'agendamento': ag,
            'etapas': etapas,
            'etapas_concluidas': etapas_concluidas,
            'total_etapas': total_etapas
        })
    
    context = {
        'agendamentos_com_etapas': agendamentos_com_etapas,
        'hoje': data_selecionada,
        'coletas_count': estatisticas['coletas'] or 0,
        'entregas_count': estatisticas['entregas'] or 0,
        'veiculos_count': veiculos_count,
        'concluidos_count': estatisticas['concluidos'] or 0,
        'peso_total': estatisticas['peso_total'] or 0,
        'is_monitor': is_monitor,
    }
    
    return render(request, 'processos_painel.html', context)


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@require_http_methods(["GET"])
def verificar_atualizacoes_processos(request):
    """
    Endpoint para verificar atualizações dos processos em tempo real
    Retorna os processos do dia com seus timestamps de atualização
    """
    data_filtro = request.GET.get('data')
    
    # Data padrão: hoje
    if data_filtro:
        try:
            data_selecionada = datetime.strptime(data_filtro, '%Y-%m-%d').date()
        except ValueError:
            data_selecionada = timezone_now().date()
    else:
        data_selecionada = timezone_now().date()
    
    # Buscar agendamentos do dia - ordenar por atualizado_em desc (mais recente primeiro), depois horario
    agendamentos = Agendamento.objects.filter(
        data_agendada=data_selecionada
    ).select_related('transportadora', 'motorista').order_by(F('atualizado_em').desc(nulls_last=True), 'horario_agendado')
    
    # Preparar dados dos processos
    processos = []
    for ag in agendamentos:
        etapas = get_etapas_ordenadas(ag)
        
        processos.append({
            'id': ag.id,
            'ordem': ag.ordem,
            'motorista': ag.motorista.nome if ag.motorista else '',
            'placa': ag.placa_veiculo,
            'tipo': ag.tipo,
            'tipo_display': ag.get_tipo_display(),
            'status_geral': ag.status_geral,
            'data': ag.data_agendada.strftime('%d/%m/%Y') if ag.data_agendada else '',
            'horario': ag.horario_agendado.strftime('%H:%M') if ag.horario_agendado else '',
            'transportadora': ag.transportadora.nome if ag.transportadora else '',
            'peso': ag.peso or 0,
            'etapas': [
                {
                    'nome': etapa['nome'],
                    'concluida': etapa['concluida'],
                    'disponivel': etapa['disponivel'],
                    'tipo': etapa.get('tipo', 'sequencial'),
                    'data': etapa['data'].isoformat() if etapa['data'] else None,
                }
                for etapa in etapas
            ],
            'atualizado_em': ag.atualizado_em.isoformat() if ag.atualizado_em else '',
            'prioridade': {
                'aguardando_chegada': 1,
                'em_checklist': 2,
                'confirmacao_armazem': 3,
                'pendente_liberacao_onda': 4,
                'processo_concluido': 5
            }.get(ag.status_geral, 6),
        })
    
    # Estatísticas
    estatisticas = agendamentos.aggregate(
        coletas=Count('id', filter=Q(tipo='coleta')),
        entregas=Count('id', filter=Q(tipo='entrega')),
        concluidos=Count('id', filter=Q(status_geral='processo_concluido')),
        peso_total=Sum('peso')
    )
    
    # Veículos = soma de coletas + entregas (cada processo é um veículo)
    veiculos_count = (estatisticas['coletas'] or 0) + (estatisticas['entregas'] or 0)
    
    return JsonResponse({
        'success': True,
        'processos': processos,
        'estatisticas': {
            'coletas': estatisticas['coletas'] or 0,
            'entregas': estatisticas['entregas'] or 0,
            'veiculos': veiculos_count,
            'concluidos': estatisticas['concluidos'] or 0,
            'peso_total': float(estatisticas['peso_total'] or 0),
        },
        'timestamp': timezone_now().isoformat(),
    })


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@require_http_methods(["GET"])
def vapid_public_key(request):
    """
    Retorna a chave pública VAPID para Web Push
    """
    from django.conf import settings
    
    vapid_public_key = getattr(settings, 'VAPID_PUBLIC_KEY', None)
    
    if not vapid_public_key:
        return JsonResponse({'error': 'VAPID keys não configuradas'}, status=500)
    
    return JsonResponse({'publicKey': vapid_public_key})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@require_http_methods(["POST"])
def registrar_push_subscription(request):
    """
    Registra a subscription do usuário para Web Push
    """
    from .models import PreferenciaNotificacaoUsuario
    import json
    
    try:
        data = json.loads(request.body)
        subscription = data.get('subscription')
        
        if not subscription:
            return JsonResponse({'success': False, 'error': 'Subscription não fornecida'})
        
        # Buscar ou criar preferências
        try:
            preferencias = PreferenciaNotificacaoUsuario.objects.get(usuario=request.user)
        except PreferenciaNotificacaoUsuario.DoesNotExist:
            preferencias = PreferenciaNotificacaoUsuario.objects.create(
                usuario=request.user,
                receber_email=True,
                receber_whatsapp=True,
                receber_navegador=True
            )
        
        # Salvar subscription
        preferencias.push_subscription = json.dumps(subscription)
        preferencias.receber_navegador = True
        preferencias.save()
        
        return JsonResponse({'success': True, 'message': 'Subscription registrada com sucesso'})
        
    except Exception as e:
        logger.error(f"Erro ao registrar subscription: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@require_http_methods(["GET"])
def notificacoes_pendentes(request):
    """
    Retorna notificações pendentes do tipo navegador para o usuário logado
    """
    from .models import NotificacaoProcesso, PreferenciaNotificacaoUsuario
    from django.utils import timezone
    
    # Verificar se usuário quer receber notificações do navegador
    try:
        preferencias = PreferenciaNotificacaoUsuario.objects.get(usuario=request.user)
        if not preferencias.receber_navegador:
            return JsonResponse({'notificacoes': []})
    except PreferenciaNotificacaoUsuario.DoesNotExist:
        return JsonResponse({'notificacoes': []})
    
    # Buscar notificações do tipo navegador dos últimos 5 minutos que ainda não foram mostradas
    cinco_minutos_atras = timezone.now() - timedelta(minutes=5)
    
    notificacoes = NotificacaoProcesso.objects.filter(
        tipo='navegador',
        enviado_com_sucesso=True,
        enviado_em__gte=cinco_minutos_atras
    ).select_related('agendamento').order_by('-enviado_em')[:10]
    
    # Filtrar apenas notificações relacionadas a processos do usuário (via grupos)
    # Por enquanto, retornar todas (pode ser melhorado depois)
    notificacoes_data = []
    for notif in notificacoes:
        notificacoes_data.append({
            'id': notif.id,
            'titulo': f'Processo {notif.agendamento.ordem}',
            'mensagem': notif.mensagem,
            'enviado_em': notif.enviado_em.isoformat(),
            'agendamento_id': notif.agendamento.id
        })
    
    return JsonResponse({'notificacoes': notificacoes_data})


def enviar_push_notification(usuario, mensagem, titulo='TLOGpainel', url=None, tag=None):
    """
    Envia notificação push para um usuário
    
    Args:
        usuario: Usuário que receberá a notificação
        mensagem: Mensagem da notificação
        titulo: Título da notificação (padrão: 'TLOGpainel')
        url: URL para abrir quando clicar na notificação (padrão: '/')
        tag: Tag para agrupar notificações similares (padrão: 'tlogpainel-notification')
    """
    from .models import PreferenciaNotificacaoUsuario
    from django.conf import settings
    import json
    
    # Importar pywebpush no início
    try:
        from pywebpush import webpush
        from pywebpush import WebPushException
    except ImportError:
        logger.error("pywebpush não instalado. Execute: pip install pywebpush")
        return False
    
    try:
        preferencias = PreferenciaNotificacaoUsuario.objects.get(usuario=usuario)
        
        if not preferencias.push_subscription or not preferencias.receber_navegador:
            return False
        
        subscription = json.loads(preferencias.push_subscription)
        
        vapid_private_key = getattr(settings, 'VAPID_PRIVATE_KEY', None)
        vapid_claims = getattr(settings, 'VAPID_CLAIMS', {
            "sub": "mailto:digitalmidia@transcamila.com.br"
        })
        
        if not vapid_private_key:
            logger.error("VAPID_PRIVATE_KEY não configurada")
            return False
        
        # URL padrão
        if url is None:
            url = '/'
        
        # Tag padrão
        if tag is None:
            tag = 'tlogpainel-notification'
        
        # Payload para Web Push - deve ser um objeto JSON simples
        payload_data = {
            'title': titulo,
            'body': mensagem,
            'icon': '/static/imagens/icone.png',
            'tag': tag,
            'url': url,  # URL para abrir quando clicar na notificação
            'data': {
                'url': url,
                'tag': tag
            }
        }
        payload = json.dumps(payload_data)
        
        webpush(
            subscription_info=subscription,
            data=payload,
            vapid_private_key=vapid_private_key,
            vapid_claims=vapid_claims
        )
        
        return True
        
    except WebPushException as e:
        logger.error(f"Erro WebPush para {usuario.username}: {e}")
        # Se subscription inválida, limpar
        if "410" in str(e) or "expired" in str(e).lower():
            try:
                preferencias.push_subscription = None
                preferencias.save()
            except:
                pass
        return False
    except Exception as e:
        logger.error(f"Erro ao enviar push notification para {usuario.username}: {e}")
        return False


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@acesso_permitido_por_aba('dashboard')
def processos_dashboard(request):
    """Dashboard analítico com dados operacionais dos agendamentos"""
    # Verificar se o usuário está no grupo "Monitores"
    is_monitor = False
    try:
        grupo_monitores = GrupoUsuario.objects.filter(nome='monitores', ativo=True).first()
        if grupo_monitores and grupo_monitores.usuarios.filter(id=request.user.id).exists():
            is_monitor = True
    except Exception:
        pass
    
    # === FILTROS ===
    periodo_selecionado = request.GET.get('periodo', 'semana')  # Padrão: Esta Semana
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    granularidade = request.GET.get('granularidade', 'dia')  # dia, mes, ano
    tipo_filtro = request.GET.get('tipo')
    transportadora_filtro = request.GET.get('transportadora')
    
    # Período padrão: mês atual (primeiro dia do mês até hoje)
    hoje = timezone_today()
    primeiro_dia_mes = hoje.replace(day=1)
    
    # Se período foi selecionado mas não há datas explícitas, calcular as datas baseado no período
    # IMPORTANTE: Verificar período ANTES de usar o padrão
    if periodo_selecionado and periodo_selecionado != 'personalizado' and not data_inicio:
        if periodo_selecionado == 'hoje':
            data_inicio = hoje.strftime('%Y-%m-%d')
            data_fim = hoje.strftime('%Y-%m-%d')
            granularidade = 'dia'
        elif periodo_selecionado == 'semana':
            inicio_semana = hoje
            # Domingo da semana
            dias_desde_domingo = inicio_semana.weekday() + 1
            inicio_semana = inicio_semana - timedelta(days=dias_desde_domingo % 7)
            data_inicio = inicio_semana.strftime('%Y-%m-%d')
            data_fim = hoje.strftime('%Y-%m-%d')
            granularidade = 'dia'
        elif periodo_selecionado == 'mes':
            primeiro_dia_mes = hoje.replace(day=1)
            data_inicio = primeiro_dia_mes.strftime('%Y-%m-%d')
            data_fim = hoje.strftime('%Y-%m-%d')
            granularidade = 'dia'
        elif periodo_selecionado == 'mes_anterior':
            primeiro_dia_mes_anterior = (hoje.replace(day=1) - timedelta(days=1)).replace(day=1)
            ultimo_dia_mes_anterior = hoje.replace(day=1) - timedelta(days=1)
            data_inicio = primeiro_dia_mes_anterior.strftime('%Y-%m-%d')
            data_fim = ultimo_dia_mes_anterior.strftime('%Y-%m-%d')
            granularidade = 'dia'
        elif periodo_selecionado == 'ano':
            primeiro_dia_ano = hoje.replace(month=1, day=1)
            data_inicio = primeiro_dia_ano.strftime('%Y-%m-%d')
            data_fim = hoje.strftime('%Y-%m-%d')
            granularidade = 'mes'
    
    # Agora processar as datas (podem ter sido definidas pelo período ou vieram do formulário)
    if data_inicio:
        try:
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        except ValueError:
            # Se período foi "hoje", usar hoje; senão usar primeiro_dia_mes como padrão
            if periodo_selecionado == 'hoje':
                data_inicio_obj = hoje
            else:
                data_inicio_obj = primeiro_dia_mes
    else:
        # Se não há data_inicio, usar padrão baseado no período selecionado
        if periodo_selecionado == 'hoje':
            data_inicio_obj = hoje
        elif periodo_selecionado == 'semana':
            # Início da semana (domingo)
            inicio_semana = hoje
            dias_desde_domingo = inicio_semana.weekday() + 1
            inicio_semana = inicio_semana - timedelta(days=dias_desde_domingo % 7)
            data_inicio_obj = inicio_semana
        else:
            data_inicio_obj = primeiro_dia_mes
    
    if data_fim:
        try:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
        except ValueError:
            data_fim_obj = hoje
    else:
        data_fim_obj = hoje
    
    # Garantir que data_inicio <= data_fim
    if data_inicio_obj > data_fim_obj:
        data_inicio_obj, data_fim_obj = data_fim_obj, data_inicio_obj
    
    # === QUERY BASE COM FILTROS DE PERÍODO ===
    agendamentos_filtrados = Agendamento.objects.filter(
        data_agendada__gte=data_inicio_obj,
        data_agendada__lte=data_fim_obj
    ).select_related('transportadora', 'motorista')
    
    # Aplicar filtros adicionais
    if tipo_filtro:
        agendamentos_filtrados = agendamentos_filtrados.filter(tipo=tipo_filtro)
    
    if transportadora_filtro:
        agendamentos_filtrados = agendamentos_filtrados.filter(transportadora_id=transportadora_filtro)

    # === KPIs ===
    total_agendamentos = agendamentos_filtrados.count()
    
    # Coletas e Entregas
    coleta_total = agendamentos_filtrados.filter(tipo='coleta').count()
    entrega_total = agendamentos_filtrados.filter(tipo='entrega').count()
    
    # Total de veículos únicos
    total_veiculos = agendamentos_filtrados.values('placa_veiculo').distinct().count()
    
    # Peso Total
    peso_total_aggr = agendamentos_filtrados.aggregate(soma=Sum('peso'))
    peso_total_kg = float(peso_total_aggr['soma'] or 0)
    
    # === MÉTRICAS DE DESEMPENHO ===
    # Processos concluídos
    processos_concluidos = agendamentos_filtrados.filter(status_geral='processo_concluido').count()
    taxa_conclusao = round((processos_concluidos / total_agendamentos * 100) if total_agendamentos > 0 else 0, 1)
    
    # Processos pendentes por etapa
    pendentes_portaria = agendamentos_filtrados.filter(status_geral='aguardando_chegada').count()
    pendentes_checklist = agendamentos_filtrados.filter(status_geral='em_checklist').count()
    pendentes_onda = agendamentos_filtrados.filter(status_geral='pendente_liberacao_onda').count()
    pendentes_armazem = agendamentos_filtrados.filter(status_geral='confirmacao_armazem').count()
    pendentes_documentos = agendamentos_filtrados.filter(status_geral='pendente_documentos').count()
    
    # Contagem por status
    status_counts = {
        'Aguardando Portaria': pendentes_portaria,
        'Em Checklist': pendentes_checklist,
        'Pendente Onda': pendentes_onda,
        'Aguardando Armazém': pendentes_armazem,
        'Pendente Documentos': pendentes_documentos,
        'Concluído': processos_concluidos
    }
    
    # Calcular tempos médios entre etapas separados por tipo (Coleta e Entrega)
    # Considerar processos que têm pelo menos uma etapa concluída no período filtrado
    tempos_coleta = {
        'portaria_checklist': [],
        'checklist_armazem': [],
        'armazem_operacao': [], # Tempo de operação (Chegada -> Saída)
        'armazem_documentos': [],
        'processo_total': []
    }
    
    tempos_entrega = {
        'portaria_armazem': [],  # Entrega pula o Checklist
        'armazem_operacao': [], # Tempo de operação (Chegada -> Saída)
        'armazem_documentos': [],
        'processo_total': []
    }
    
    # Buscar processos que têm pelo menos Portaria concluída E que foram agendados no período
    # OU que têm alguma etapa concluída no período (para pegar processos em andamento)
    processos_parciais = Agendamento.objects.filter(
        Q(
            # Processos agendados no período E têm pelo menos Portaria
            data_agendada__gte=data_inicio_obj,
            data_agendada__lte=data_fim_obj
        ) | Q(
            # OU processos que têm alguma etapa concluída no período
            portaria_liberacao__date__gte=data_inicio_obj,
            portaria_liberacao__date__lte=data_fim_obj
        ) | Q(
            checklist_data__date__gte=data_inicio_obj,
            checklist_data__date__lte=data_fim_obj
        ) | Q(
            armazem_chegada__date__gte=data_inicio_obj,
            armazem_chegada__date__lte=data_fim_obj
        ) | Q(
            documentos_liberacao__date__gte=data_inicio_obj,
            documentos_liberacao__date__lte=data_fim_obj
        )
    ).filter(
        # Mas precisa ter pelo menos Portaria concluída para calcular tempos
        portaria_liberacao__isnull=False
    )
    
    # Aplicar filtros adicionais se existirem
    if tipo_filtro:
        processos_parciais = processos_parciais.filter(tipo=tipo_filtro)
    if transportadora_filtro:
        processos_parciais = processos_parciais.filter(transportadora_id=transportadora_filtro)
    
    processos_parciais = processos_parciais.select_related('transportadora', 'motorista').distinct()
    
    for ag in processos_parciais:
        # Selecionar o dicionário correto baseado no tipo
        if ag.tipo == 'coleta':
            # COLETA: Portaria → Checklist → Armazém → Documentos
            
            # Tempo Portaria -> Checklist (precisa ter ambas concluídas)
            if ag.portaria_liberacao and ag.checklist_data:
                diferenca = ag.checklist_data - ag.portaria_liberacao
                tempo = diferenca.total_seconds() / 3600.0  # em horas decimais
                if tempo >= 0:  # Permitir 0 também (processos muito rápidos)
                    tempos_coleta['portaria_checklist'].append(tempo)
            
            # Tempo Checklist -> Armazém (precisa ter ambas concluídas)
            if ag.checklist_data and ag.armazem_chegada:
                diferenca = ag.armazem_chegada - ag.checklist_data
                tempo = diferenca.total_seconds() / 3600.0
                if tempo >= 0:
                    tempos_coleta['checklist_armazem'].append(tempo)
            
            # Tempo Operação Armazém (Chegada -> Saída)
            if ag.armazem_chegada and ag.armazem_saida:
                diferenca = ag.armazem_saida - ag.armazem_chegada
                tempo = diferenca.total_seconds() / 3600.0
                if tempo >= 0:
                    tempos_coleta['armazem_operacao'].append(tempo)
            
            # Tempo Armazém -> Documentos (Início Operação Armazém -> Liberação Documentos)
            if ag.armazem_chegada and ag.documentos_liberacao:
                diferenca = ag.documentos_liberacao - ag.armazem_chegada
                tempo = diferenca.total_seconds() / 3600.0  # em horas decimais
                if tempo >= 0:
                    tempos_coleta['armazem_documentos'].append(tempo)
            
            # Tempo total do processo: desde Portaria até a última etapa concluída
            if ag.portaria_liberacao:
                # Determinar qual foi a última etapa concluída
                ultima_etapa_data = ag.portaria_liberacao  # Inicializar com portaria
                
                if ag.documentos_liberacao:
                    ultima_etapa_data = ag.documentos_liberacao
                elif ag.armazem_chegada:
                    ultima_etapa_data = ag.armazem_chegada
                elif ag.checklist_data:
                    ultima_etapa_data = ag.checklist_data
                
                # Só adicionar se houve progresso (saiu da portaria)
                if ultima_etapa_data != ag.portaria_liberacao:
                    diferenca = ultima_etapa_data - ag.portaria_liberacao
                    tempo = diferenca.total_seconds() / 3600.0  # em horas decimais
                    if tempo >= 0:
                        tempos_coleta['processo_total'].append(tempo)
        
        elif ag.tipo == 'entrega':
            # ENTREGA: Portaria → Armazém → Documentos (pula Checklist)
            
            # Tempo Portaria -> Armazém (precisa ter ambas concluídas)
            if ag.portaria_liberacao and ag.armazem_chegada:
                diferenca = ag.armazem_chegada - ag.portaria_liberacao
                tempo = diferenca.total_seconds() / 3600.0  # em horas decimais
                if tempo >= 0:
                    tempos_entrega['portaria_armazem'].append(tempo)
            
            # Tempo Operação Armazém (Chegada -> Saída)
            if ag.armazem_chegada and ag.armazem_saida:
                diferenca = ag.armazem_saida - ag.armazem_chegada
                tempo = diferenca.total_seconds() / 3600.0
                if tempo >= 0:
                    tempos_entrega['armazem_operacao'].append(tempo)
            
            # Tempo Armazém -> Documentos (Início Operação Armazém -> Liberação Documentos)
            if ag.armazem_chegada and ag.documentos_liberacao:
                diferenca = ag.documentos_liberacao - ag.armazem_chegada
                tempo = diferenca.total_seconds() / 3600.0  # em horas decimais
                if tempo >= 0:
                    tempos_entrega['armazem_documentos'].append(tempo)
            
            # Tempo total do processo: desde Portaria até a última etapa concluída
            if ag.portaria_liberacao:
                # Determinar qual foi a última etapa concluída
                ultima_etapa_data = ag.portaria_liberacao  # Inicializar com portaria
                
                if ag.documentos_liberacao:
                    ultima_etapa_data = ag.documentos_liberacao
                elif ag.armazem_chegada:
                    ultima_etapa_data = ag.armazem_chegada
                
                # Só adicionar se houve progresso (saiu da portaria)
                if ultima_etapa_data != ag.portaria_liberacao:
                    diferenca = ultima_etapa_data - ag.portaria_liberacao
                    tempo = diferenca.total_seconds() / 3600.0  # em horas decimais
                    if tempo >= 0:
                        tempos_entrega['processo_total'].append(tempo)
    
    # Função auxiliar para converter horas decimais em formato HH:MM:SS
    def formatar_tempo_medio(tempo_horas):
        if tempo_horas == 0 or tempo_horas is None:
            return {'texto': '-'}
        # Converter horas decimais em segundos totais (usar round para evitar problemas de precisão)
        segundos_totais = round(tempo_horas * 3600)
        # Calcular horas, minutos e segundos
        horas = int(segundos_totais // 3600)
        minutos = int((segundos_totais % 3600) // 60)
        segundos = int(segundos_totais % 60)
        # Formatar como HH:MM:SS
        texto = f"{horas:02d}:{minutos:02d}:{segundos:02d}"
        return {'texto': texto}
    
    # Função para calcular médias de COLETA (com todas as etapas)
    def calcular_medias_coleta(tempos_dict, prefixo_log=''):
        resultados = {}
        
        # Portaria -> Checklist
        if tempos_dict['portaria_checklist']:
            soma = sum(tempos_dict['portaria_checklist'])
            qtd = len(tempos_dict['portaria_checklist'])
            media_h = soma / qtd
            if prefixo_log:
                logger.debug(f"{prefixo_log} Portaria->Checklist: {qtd} processos, soma={soma:.4f}h, média={media_h:.4f}h")
            resultados['portaria_checklist'] = formatar_tempo_medio(media_h)
        else:
            resultados['portaria_checklist'] = formatar_tempo_medio(0)
        
        # Checklist -> Armazém
        if tempos_dict['checklist_armazem']:
            soma = sum(tempos_dict['checklist_armazem'])
            qtd = len(tempos_dict['checklist_armazem'])
            media_h = soma / qtd
            if prefixo_log:
                logger.debug(f"{prefixo_log} Checklist->Armazém: {qtd} processos, soma={soma:.4f}h, média={media_h:.4f}h")
            resultados['checklist_armazem'] = formatar_tempo_medio(media_h)
        else:
            resultados['checklist_armazem'] = formatar_tempo_medio(0)
        
        # Operação Armazém
        if tempos_dict.get('armazem_operacao'):
            soma = sum(tempos_dict['armazem_operacao'])
            qtd = len(tempos_dict['armazem_operacao'])
            media_h = soma / qtd
            if prefixo_log:
                logger.debug(f"{prefixo_log} Operação Armazém: {qtd} processos, soma={soma:.4f}h, média={media_h:.4f}h")
            resultados['armazem_operacao'] = formatar_tempo_medio(media_h)
        else:
            resultados['armazem_operacao'] = formatar_tempo_medio(0)

        # Armazém -> Documentos
        if tempos_dict['armazem_documentos']:
            soma = sum(tempos_dict['armazem_documentos'])
            qtd = len(tempos_dict['armazem_documentos'])
            media_h = soma / qtd
            if prefixo_log:
                logger.debug(f"{prefixo_log} Armazém->Documentos: {qtd} processos, soma={soma:.4f}h, média={media_h:.4f}h")
            resultados['armazem_documentos'] = formatar_tempo_medio(media_h)
        else:
            resultados['armazem_documentos'] = formatar_tempo_medio(0)
        
        # Processo Total
        if tempos_dict['processo_total']:
            soma = sum(tempos_dict['processo_total'])
            qtd = len(tempos_dict['processo_total'])
            media_h = soma / qtd
            if prefixo_log:
                logger.debug(f"{prefixo_log} Processo Total: {qtd} processos, soma={soma:.4f}h, média={media_h:.4f}h")
            resultados['processo_total'] = formatar_tempo_medio(media_h)
        else:
            resultados['processo_total'] = formatar_tempo_medio(0)
        
        return resultados
    
    # Função para calcular médias de ENTREGA (pula checklist)
    def calcular_medias_entrega(tempos_dict, prefixo_log=''):
        resultados = {}
        
        # Portaria -> Armazém (delivery pula checklist)
        if tempos_dict['portaria_armazem']:
            soma = sum(tempos_dict['portaria_armazem'])
            qtd = len(tempos_dict['portaria_armazem'])
            media_h = soma / qtd
            if prefixo_log:
                logger.debug(f"{prefixo_log} Portaria->Armazém: {qtd} processos, soma={soma:.4f}h, média={media_h:.4f}h")
            resultados['portaria_armazem'] = formatar_tempo_medio(media_h)
        else:
            resultados['portaria_armazem'] = formatar_tempo_medio(0)
        
        # Operação Armazém
        if tempos_dict.get('armazem_operacao'):
            soma = sum(tempos_dict['armazem_operacao'])
            qtd = len(tempos_dict['armazem_operacao'])
            media_h = soma / qtd
            resultados['armazem_operacao'] = formatar_tempo_medio(media_h)
        else:
            resultados['armazem_operacao'] = formatar_tempo_medio(0)
        
        # Armazém -> Documentos
        if tempos_dict['armazem_documentos']:
            soma = sum(tempos_dict['armazem_documentos'])
            qtd = len(tempos_dict['armazem_documentos'])
            media_h = soma / qtd
            if prefixo_log:
                logger.debug(f"{prefixo_log} Armazém->Documentos: {qtd} processos, soma={soma:.4f}h, média={media_h:.4f}h")
            resultados['armazem_documentos'] = formatar_tempo_medio(media_h)
        else:
            resultados['armazem_documentos'] = formatar_tempo_medio(0)
        
        # Processo Total
        if tempos_dict['processo_total']:
            soma = sum(tempos_dict['processo_total'])
            qtd = len(tempos_dict['processo_total'])
            media_h = soma / qtd
            if prefixo_log:
                logger.debug(f"{prefixo_log} Processo Total: {qtd} processos, soma={soma:.4f}h, média={media_h:.4f}h")
            resultados['processo_total'] = formatar_tempo_medio(media_h)
        else:
            resultados['processo_total'] = formatar_tempo_medio(0)
        
        return resultados
    
    # Calcular médias separadas para Coleta e Entrega
    tempos_medios_coleta = calcular_medias_coleta(tempos_coleta, '[COLETA]')
    tempos_medios_entrega = calcular_medias_entrega(tempos_entrega, '[ENTREGA]')
    
    # Processos concluídos hoje
    processos_concluidos_hoje = agendamentos_filtrados.filter(
        status_geral='processo_concluido',
        documentos_liberacao__date=hoje
    ).count()

    # === Gráfico 1: Agendamentos por Período (com granularidade) ===
    # Buscar todos os agendamentos com data_agendada (mais eficiente para SQLite)
    agendamentos_list = list(agendamentos_filtrados.values('data_agendada', 'id'))
    
    labels_periodo = []
    valores_periodo = []
    agrupados_dict = {}
    
    # Agrupar em Python (compatível com SQLite)
    for agendamento in agendamentos_list:
        data = agendamento['data_agendada']
        
        if granularidade == 'dia':
            key = data
        elif granularidade == 'mes':
            # Usar primeiro dia do mês como chave
            key = data.replace(day=1)
        elif granularidade == 'ano':
            # Usar primeiro dia do ano como chave
            key = data.replace(month=1, day=1)
        
        agrupados_dict[key] = agrupados_dict.get(key, 0) + 1
    
    # Preencher todos os períodos
    if granularidade == 'dia':
        current_date = data_inicio_obj
        while current_date <= data_fim_obj:
            total = agrupados_dict.get(current_date, 0)
            labels_periodo.append(current_date.strftime("%d/%m"))
            valores_periodo.append(total)
            current_date += timedelta(days=1)
            
    elif granularidade == 'mes':
        current_date = data_inicio_obj.replace(day=1)
        fim_date = data_fim_obj.replace(day=1)
        while current_date <= fim_date:
            total = agrupados_dict.get(current_date, 0)
            labels_periodo.append(current_date.strftime("%m/%Y"))
            valores_periodo.append(total)
            # Próximo mês
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)
                
    elif granularidade == 'ano':
        current_year = data_inicio_obj.year
        while current_year <= data_fim_obj.year:
            key = datetime(current_year, 1, 1).date()
            total = agrupados_dict.get(key, 0)
            labels_periodo.append(str(current_year))
            valores_periodo.append(total)
            current_year += 1

    dados_periodo = {
        'labels': labels_periodo,
        'valores': valores_periodo
    }

    # === Gráfico 2: Por Transportadora ===
    top_transp = agendamentos_filtrados.values('transportadora__nome')\
        .annotate(total=Count('id')).order_by('-total')[:10]

    transportadoras_labels = []
    transportadoras_valores = []
    
    for item in top_transp:
        nome = item['transportadora__nome'] or 'Sem Transportadora'
        transportadoras_labels.append(nome)
        transportadoras_valores.append(item['total'])
    
    if not transportadoras_labels:
        transportadoras_labels = ['Sem dados']
        transportadoras_valores = [0]

    dados_transportadoras = {
        'labels': transportadoras_labels,
        'valores': transportadoras_valores
    }

    # === Dados para gráfico de status ===
    dados_status = {
        'labels': list(status_counts.keys()),
        'valores': list(status_counts.values())
    }

    # Transportadoras para o filtro
    transportadoras = Transportadora.objects.all().order_by('nome')
    
    context = {
        'agora': timezone_now(),
        'data_inicio': data_inicio_obj,
        'data_fim': data_fim_obj,
        'granularidade': granularidade,
        'total_agendamentos': total_agendamentos,
        'coleta_total': coleta_total,
        'entrega_total': entrega_total,
        'total_veiculos': total_veiculos,
        'peso_total_kg': peso_total_kg,
        'processos_concluidos': processos_concluidos,
        'taxa_conclusao': taxa_conclusao,
        'pendentes_portaria': pendentes_portaria,
        'pendentes_checklist': pendentes_checklist,
        'pendentes_onda': pendentes_onda,
        'pendentes_armazem': pendentes_armazem,
        'pendentes_documentos': pendentes_documentos,
        'tempo_medio_coleta_portaria_checklist': tempos_medios_coleta['portaria_checklist'],
        'tempo_medio_coleta_checklist_armazem': tempos_medios_coleta['checklist_armazem'],
        'tempo_medio_coleta_armazem_operacao': tempos_medios_coleta['armazem_operacao'],
        'tempo_medio_coleta_armazem_documentos': tempos_medios_coleta['armazem_documentos'],
        'tempo_medio_coleta_total': tempos_medios_coleta['processo_total'],
        'tempo_medio_entrega_portaria_armazem': tempos_medios_entrega['portaria_armazem'],
        'tempo_medio_entrega_armazem_operacao': tempos_medios_entrega['armazem_operacao'],
        'tempo_medio_entrega_armazem_documentos': tempos_medios_entrega['armazem_documentos'],
        'tempo_medio_entrega_total': tempos_medios_entrega['processo_total'],
        'processos_concluidos_hoje': processos_concluidos_hoje,
        'dados_periodo': json.dumps(dados_periodo, ensure_ascii=False),
        'dados_transportadoras': json.dumps(dados_transportadoras, ensure_ascii=False),
        'dados_status': json.dumps(dados_status, ensure_ascii=False),
        'transportadoras': transportadoras,
        'tipo_filtro': tipo_filtro,
        'transportadora_filtro': transportadora_filtro,
        'periodo_selecionado': periodo_selecionado,
        'is_monitor': is_monitor,
    }

    return render(request, 'processos_dashboard.html', context)


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@require_http_methods(["GET"])
def notificacoes_pendentes(request):
    """
    Retorna notificações pendentes do tipo navegador para o usuário logado
    """
    from .models import NotificacaoProcesso, PreferenciaNotificacaoUsuario
    
    # Verificar se usuário quer receber notificações do navegador
    try:
        preferencias = PreferenciaNotificacaoUsuario.objects.get(usuario=request.user)
        if not preferencias.receber_navegador:
            return JsonResponse({'notificacoes': []})
    except PreferenciaNotificacaoUsuario.DoesNotExist:
        return JsonResponse({'notificacoes': []})
    
    # Buscar notificações do tipo navegador dos últimos 5 minutos
    cinco_minutos_atras = django_timezone.now() - timedelta(minutes=5)
    
    notificacoes = NotificacaoProcesso.objects.filter(
        tipo='navegador',
        enviado_com_sucesso=True,
        enviado_em__gte=cinco_minutos_atras
    ).select_related('agendamento').order_by('-enviado_em')[:10]
    
    # Preparar dados das notificações
    notificacoes_data = []
    ids_mostrados = set()
    
    for notif in notificacoes:
        # Evitar duplicatas
        if notif.id in ids_mostrados:
            continue
        ids_mostrados.add(notif.id)
        
        notificacoes_data.append({
            'id': notif.id,
            'titulo': f'Processo {notif.agendamento.ordem}',
            'mensagem': notif.mensagem,
            'enviado_em': notif.enviado_em.isoformat(),
            'agendamento_id': notif.agendamento.id
        })
    
    return JsonResponse({'notificacoes': notificacoes_data})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@require_http_methods(["POST"])
def cadastrar_motorista_rapido(request):
    """
    View para cadastrar motorista rapidamente via AJAX
    """
    try:
        import json
        data = json.loads(request.body)
        nome = data.get('nome', '').strip().upper()
        telefone = data.get('telefone', '').strip()
        
        if not nome:
            return JsonResponse({'success': False, 'error': 'Nome é obrigatório'})
        
        # Verificar se já existe motorista com mesmo nome (case-insensitive)
        motorista_existente = Motorista.objects.filter(nome__iexact=nome).first()
        if motorista_existente:
            return JsonResponse({
                'success': False,
                'error': f'Motorista "{motorista_existente.nome}" já está cadastrado',
                'motorista_id': motorista_existente.id,
                'motorista_nome': motorista_existente.nome
            })
        
        # Criar novo motorista
        motorista = Motorista.objects.create(
            nome=nome,
            telefone=telefone if telefone else None
        )
        
        return JsonResponse({
            'success': True,
            'motorista_id': motorista.id,
            'message': 'Motorista cadastrado com sucesso'
        })
        
    except Exception as e:
        logger.error(f"Erro ao cadastrar motorista rápido: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erro: {str(e)}'})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
@require_http_methods(["POST"])
def cadastrar_transportadora_rapida(request):
    """
    View para cadastrar transportadora rapidamente via AJAX
    """
    try:
        import json
        data = json.loads(request.body)
        nome = data.get('nome', '').strip().upper()
        cnpj = data.get('cnpj', '').strip()
        telefone = data.get('telefone', '').strip()
        
        if not nome:
            return JsonResponse({'success': False, 'error': 'Nome é obrigatório'})
        
        # Verificar se já existe transportadora com mesmo nome (case-insensitive)
        transportadora_existente = Transportadora.objects.filter(nome__iexact=nome).first()
        if transportadora_existente:
            return JsonResponse({
                'success': False,
                'error': f'Transportadora "{transportadora_existente.nome}" já está cadastrada',
                'transportadora_id': transportadora_existente.id,
                'transportadora_nome': transportadora_existente.nome
            })
        
        # Criar nova transportadora
        transportadora = Transportadora.objects.create(
            nome=nome,
            cnpj=cnpj if cnpj else None,
            telefone=telefone if telefone else None
        )
        
        return JsonResponse({
            'success': True,
            'transportadora_id': transportadora.id,
            'message': 'Transportadora cadastrada com sucesso'
        })
        
    except Exception as e:
        logger.error(f"Erro ao cadastrar transportadora rápida: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erro: {str(e)}'})


@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
def atualizar_nomes_para_maiusculas(request):
    """
    View administrativa para atualizar todos os nomes para maiúsculas
    """
    try:
        from django.db import transaction
        
        with transaction.atomic():
            motoristas_atualizados = 0
            for motorista in Motorista.objects.all():
                nome_antigo = motorista.nome
                nome_novo = motorista.nome.strip().upper()
                if nome_antigo != nome_novo:
                    motorista.nome = nome_novo
                    motorista.save()
                    motoristas_atualizados += 1
            
            transportadoras_atualizadas = 0
            for transportadora in Transportadora.objects.all():
                nome_antigo = transportadora.nome
                nome_novo = transportadora.nome.strip().upper()
                if nome_antigo != nome_novo:
                    transportadora.nome = nome_novo
                    transportadora.save()
                    transportadoras_atualizadas += 1
        
        return JsonResponse({
            'success': True,
            'message': f'{motoristas_atualizados} motorista(s) e {transportadoras_atualizadas} transportadora(s) atualizado(s) para maiúsculas'
        })
    except Exception as e:
        logger.error(f"Erro ao atualizar nomes: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erro: {str(e)}'})

# ==========================================================
# SMART UPDATE SYSTEM VIEWS
# ==========================================================

@login_required
def verificar_atualizacoes(request):
    """
    Endpoint leve para verificação de atualizações (Polling Inteligente).
    Recebe um timestamp (isoformat) e retorna se houve atualização.
    """
    tela = request.GET.get('tela')
    ultimo_timestamp_client = request.GET.get('timestamp')
    
    if not tela:
        return JsonResponse({'update': False, 'erro': 'Tela não informada'})
        
    try:
        from .models import ControleAtualizacao
        from django.utils.dateparse import parse_datetime
        from django.utils import timezone
        
        # Obter timestamp da última atualização global para esta tela
        controle = ControleAtualizacao.objects.filter(tela=tela).first()
        
        if not controle:
            return JsonResponse({'update': False, 'timestamp': timezone_now().isoformat()})
            
        server_timestamp = controle.ultima_atualizacao
        
        should_update = False
        
        if ultimo_timestamp_client and ultimo_timestamp_client != 'null':
            try:
                client_dt = parse_datetime(ultimo_timestamp_client)
                if client_dt:
                   # Normalizar para garantir comparação justa (evitar erro can't compare offset-naive and offset-aware)
                   if timezone.is_aware(server_timestamp) and timezone.is_naive(client_dt):
                       client_dt = timezone.make_aware(client_dt)
                   elif timezone.is_naive(server_timestamp) and timezone.is_aware(client_dt):
                       server_timestamp = timezone.make_aware(server_timestamp)
                       
                   if server_timestamp > client_dt:
                       should_update = True
            except Exception as e:
                logger.warning(f"Erro ao comparar timestamps ({tela}): {e}")
                should_update = True # Se erro na comparação, força atualizar por segurança
        else:
            # Se cliente não mandou timestamp, assume que precisa atualizar
            should_update = False # Primeira carga já tem dados
            
        return JsonResponse({
            'update': should_update,
            'timestamp': server_timestamp.isoformat()
        })
        
    except Exception as e:
        logger.error(f"Erro em verificar_atualizacoes: {e}")
        return JsonResponse({'update': False})

@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
def portaria_tabela(request):
    """Retorna apenas o partial HTML da tabela da portaria"""
    hoje = timezone_now().date()
    # ESTRATÉGIA NUCLEAR: Separação Explicita de Listas
    agendados = Agendamento.objects.filter(data_agendada=hoje, portaria_liberacao__isnull=True).select_related('transportadora', 'motorista').order_by('horario_agendado')
    
    liberados_qs = Agendamento.objects.filter(data_agendada=hoje, portaria_liberacao__isnull=False).select_related('transportadora', 'motorista', 'portaria_liberado_por')
    todos_liberados = list(liberados_qs)
    
    sem_armazem = [x for x in todos_liberados if not x.portaria_chegada_armazem]
    sem_armazem.sort(key=lambda x: x.portaria_liberacao.timestamp() if x.portaria_liberacao else 0, reverse=True)
    
    com_armazem = [x for x in todos_liberados if x.portaria_chegada_armazem]
    com_armazem.sort(key=lambda x: x.portaria_chegada_armazem.timestamp())
    
    liberados_final = sem_armazem + com_armazem
    
    # logger.info(f"PORTARIA TABELA (LEGACY) EXECUTADA - {len(liberados_final)} items")

    return render(request, 'partials/_tabela_portaria.html', {'agendamentos': agendados, 'liberados': liberados_final, 'agora': timezone_now()})

@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
def checklist_tabela(request):
    """Retorna apenas o partial HTML da tabela de checklist"""
    hoje = timezone_now().date()
    # Pendentes: liberados na portaria mas sem checklist (APENAS COLETA)
    pendentes = Agendamento.objects.filter(data_agendada=hoje, portaria_liberacao__isnull=False, checklist_data__isnull=True, tipo='coleta').select_related('transportadora', 'motorista', 'portaria_liberado_por').order_by('portaria_liberacao')
    # Concluidos: já preencheram checklist
    concluidos = Agendamento.objects.filter(data_agendada=hoje, checklist_data__isnull=False).select_related('transportadora', 'motorista', 'checklist_preenchido_por').order_by('-checklist_data')
    return render(request, 'partials/_tabela_checklist.html', {'pendentes': pendentes, 'concluidos': concluidos})

@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
def onda_tabela(request):
    """Retorna apenas o partial HTML da tabela de onda"""
    hoje = timezone_now().date()
    # Pendentes: agendamentos sem liberação de onda (APENAS COLETA) - Align with liberacao_onda view
    pendentes = Agendamento.objects.filter(data_agendada=hoje, onda_liberacao__isnull=True, tipo='coleta').select_related('transportadora', 'motorista').order_by('horario_agendado')
    concluidos = Agendamento.objects.filter(data_agendada=hoje, onda_liberacao__isnull=False).select_related('transportadora', 'motorista', 'onda_liberado_por').order_by('-onda_liberacao')
    return render(request, 'partials/_tabela_onda.html', {'pendentes': pendentes, 'concluidos': concluidos})

@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
def armazem_tabela(request):
    """Retorna apenas o partial HTML da tabela de armazem"""
    hoje = timezone_now().date()
    # Lógica combinada do armazém (CORRIGIDA)
    # PENDENTES: Inclui itens AGUARDANDO (armazem_chegada=None) OU EM OPERAÇÃO (armazem_chegada!=None e armazem_saida=None)
    pendentes = Agendamento.objects.filter(
        data_agendada=hoje, 
        armazem_saida__isnull=True
    ).filter(
        (Q(tipo='coleta') & Q(checklist_data__isnull=False)) | 
        (Q(tipo='entrega') & Q(portaria_liberacao__isnull=False))
    ).select_related('transportadora', 'motorista').order_by('horario_agendado')
    
    # CONCLUÍDOS: Apenas quando tem SAÍDA registrada
    concluidos = Agendamento.objects.filter(
        data_agendada=hoje, 
        armazem_saida__isnull=False
    ).select_related('transportadora', 'motorista', 'armazem_confirmado_por', 'armazem_saida_por').order_by('-armazem_saida')
    
    return render(request, 'partials/_tabela_armazem.html', {'pendentes': pendentes, 'concluidos': concluidos})

@login_required
@acesso_permitido_apenas_para_filial('rondonopolis')
def documentos_tabela(request):
    """Retorna apenas o partial HTML da tabela de documentos"""
    hoje = timezone_now().date()
    # Regra: Entrou no armazém, mas não liberou documentos
    pendentes = Agendamento.objects.filter(data_agendada=hoje, armazem_chegada__isnull=False, documentos_liberacao__isnull=True).select_related('transportadora', 'motorista').order_by('armazem_chegada')
    concluidos = Agendamento.objects.filter(data_agendada=hoje, documentos_liberacao__isnull=False).select_related('transportadora', 'motorista', 'documentos_liberado_por').order_by('-documentos_liberacao')
    return render(request, 'partials/_tabela_documentos.html', {'pendentes': pendentes, 'concluidos': concluidos})

@login_required
@require_http_methods(["POST"])
def confirmar_chegada_armazem_portaria(request):
    """
    Permite que a Portaria confirme que o veículo chegou no Armazém.
    Isso é diferente do processo de entrada oficial do Armazém.
    Apenas registra um timestamp de controle.
    """
    import json
    from django.shortcuts import get_object_or_404
    from django.utils import timezone
    from .models import Agendamento
    
    agendamento_id = request.POST.get('agendamento_id')
    
    if not agendamento_id:
        try:
            data = json.loads(request.body)
            agendamento_id = data.get('agendamento_id')
        except:
            pass
            
    if not agendamento_id:
        return JsonResponse({'success': False, 'error': 'ID do agendamento não fornecido'})

    try:
        ag = get_object_or_404(Agendamento, id=agendamento_id)
        
        # Já foi confirmado?
        if ag.portaria_chegada_armazem:
            return JsonResponse({'success': False, 'error': 'Chegada no armazém já foi confirmada anteriormente.'})

        # Prepare timestamp
        agora_rdn = timezone.now()
        
        # Check for manual time override
        horario_manual = request.POST.get('horario')
        if not horario_manual:
             # Try JSON body fallback
             try:
                 data = json.loads(request.body)
                 horario_manual = data.get('horario')
             except:
                 pass

        if horario_manual:
             try:
                 # Helper to combine current date with manual time
                 # Parse HH:MM
                 hora, minuto = map(int, horario_manual.split(':'))
                 # We want today's date but with manual time
                 agora_rdn = agora_rdn.replace(hour=hora, minute=minuto, second=0, microsecond=0)
                 
                 # Safety check: if time is drastically in future (e.g. user typed 23:59 at 8am), 
                 # maybe warn? But usually we trust user input. 
                 # However, if user enters a time that has passed YESTERDAY (e.g. 23:30 when it's 00:10),
                 # we might want to support "yesterday" logic? 
                 # For now, KEEP SIMPLE: Always today's date. User requested simple manual entry.
             except (ValueError, TypeError):
                 pass # Fallback to current time if invalid format

        ag.portaria_chegada_armazem = agora_rdn
        ag.portaria_chegada_armazem_por = request.user
        ag.save(update_fields=['portaria_chegada_armazem', 'portaria_chegada_armazem_por'])
        
        if hasattr(request, 'user'):
             print(f"PORTARIA - Chegada no armazém confirmada por {request.user} - Placa: {ag.placa_veiculo}")
        
        return JsonResponse({
            'success': True, 
            'message': 'Chegada no armazém confirmada com sucesso!'
        })
        
    except Exception as e:
        print(f"Erro ao confirmar chegada no armazém pela portaria (ID {agendamento_id}): {e}")
        return JsonResponse({'success': False, 'error': f'Erro ao confirmar: {str(e)}'})


@login_required
@require_http_methods(["POST"])
def armazem_registrar_saida(request):
    """
    Registra a saída do veículo do armazém.
    Usa os campos: armazem_saida e armazem_saida_por (ou o que for confirmar a saída)
    """
    ids = request.POST.getlist('ids[]')
    if not ids:
        return JsonResponse({'success': False, 'error': 'Nenhum veículo selecionado'})

    sucessos = 0
    for ag_id in ids:
        try:
            with transaction.atomic():
                ag = Agendamento.objects.select_for_update().get(id=ag_id)

                # Já saiu? Pula
                if ag.armazem_saida is not None:
                    continue

                # Registra saída
                agora_rdn = timezone_now()
                ag.armazem_saida = agora_rdn
                ag.armazem_saida_por = request.user  # Registrar quem finalizou a operação
                
                # Salvar observação se fornecida
                observacao = request.POST.get('observacao', '').strip()
                if observacao:
                    ag.armazem_saida_observacao = observacao
                
                ag.atualizar_status_geral()
                ag.save()

                enviar_notificacao_etapa(ag, 'armazem_saida', request.user)

                sucessos += 1
                logger.info(f"ARMAZÉM - Saída registrada por {request.user} - Placa: {ag.placa_veiculo}")

        except Agendamento.DoesNotExist:
            continue
        except Exception as e:
            logger.error(f"Erro ao registrar saída do armazém (ID {ag_id}): {e}")

    if sucessos > 0:
        return JsonResponse({
            'success': True,
            'message': f'Saída do armazém registrada em {sucessos} veículo(s)!'
        })
    else:
        return JsonResponse({
            'success': False,
            'error': 'Nenhum veículo foi registrado (já tinham saído ou erro)'
        })
